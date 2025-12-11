"""
主窗口（完整功能版）
完整迁移自原始main.py，保留所有功能和UI布局
"""
import sys
import os

# 路径设置
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# 标准库
import ctypes
import json
import re
import threading
import time
import traceback
import weakref
import serial
import platform
import math
from collections import deque
import csv
from datetime import datetime
import numpy as np

# Windows特定导入
if platform.system() == 'Windows':
    from ctypes import windll

# 可选依赖
try:
    import nidaqmx
    from nidaqmx.constants import TerminalConfiguration
    import pyqtgraph as pg
    NIDAQMX_AVAILABLE = True
except ImportError:
    NIDAQMX_AVAILABLE = False
    nidaqmx = None
    pg = None

# PySide6
from serial.tools import list_ports
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QDialog, QFrame, QLabel, QPushButton, QTextEdit,
    QComboBox, QLineEdit, QGridLayout, QVBoxLayout, QHBoxLayout,
    QTreeWidgetItem, QStatusBar, QInputDialog, QMessageBox, QRadioButton,
    QButtonGroup, QCheckBox, QTabWidget, QGroupBox, QSizePolicy, 
    QHeaderView, QTableWidget, QTableWidgetItem, QFormLayout,
    QFileDialog, QSplitter, QSpinBox, QDoubleSpinBox, QScrollArea, QTreeWidget,
    QMenu
)
from PySide6.QtCore import Qt, Signal, QTimer, QPointF, QMargins
from PySide6.QtGui import QFont, QTextCursor, QIcon, QColor, QPainter, QPen, QBrush, QDoubleValidator
from PySide6.QtCharts import QChartView

# 导入重构后的组件
from src.config.constants import MACOS_STYLE
from src.config.settings import SettingsManager
from src.ui.widgets import IOSSwitch, MotorCircle, AnalysisChart, DragDropTreeWidget, PIDAnalysisChart, PIDStatsPanel, PIDOptimizerPanel
from src.ui.dialogs.motor_step_config import MotorStepConfig
from src.hardware.daq_thread import DAQThread
from src.hardware.serial_reader import SerialReader
from src.core.automation_engine import AutomationThread
from src.core.pid_analyzer import PIDAnalyzer, PIDStatus
from src.core.pid_optimizer import PatternSearchOptimizer, PIDParams, TestResult

class MotorControlApp(QMainWindow):
    log_signal = Signal(str)
    angle_update = Signal(dict)

    def __init__(self):
        super().__init__()

        # --- 新增：设置文件路径 ---
        self.settings_file = "data/settings.json"
        
        # --- 初始化设置管理器 ---
        self.settings_manager = SettingsManager(self.settings_file)
        self.settings_manager.load()

        # --- 初始化光谱仪相关变量 ---
        if not NIDAQMX_AVAILABLE:
            QMessageBox.critical(self, "依赖库缺失",
                                 "未找到 nidaqmx 或 pyqtgraph 库。\n请运行 'pip install nidaqmx pyqtgraph scipy' 进行安装。\n光谱仪功能将不可用。")

        self._spectro_init_vars()

        # --- 原有初始化 ---
        # 保留日志信号连接（角度更新信号不打印，避免刷屏）
        self.log_signal.connect(lambda x: print(f"Log: {x}"))
        self.resize_timer = QTimer()
        self.resize_timer.setSingleShot(True)
        self.resize_timer.timeout.connect(self.handle_resize)
        self.serial_port = None
        self.automation_steps = []
        self.current_step = 0
        self.running = False
        self.serial_lock = threading.Lock()
        self.loop_count = 0
        # 使用重构后的PresetManager
        from src.core.preset_manager import PresetManager as PM
        self._preset_manager = PM()
        self.presets = self._preset_manager.presets
        self.automation_thread = None
        self.log_signal.connect(self._log_impl)
        self.init_ui()
        self.update_preset_combos()
        self.setStyleSheet(MACOS_STYLE)
        self.setMinimumSize(1280, 960)
        self.resize(1280, 960)
        try:
            self.setWindowIcon(QIcon('resources/icons/meow.ico'))
        except:
            pass
        self.angle_update.connect(self.update_angles)
        self.expected_changes = {}  # 记录每个电机理论转动角度
        self.last_angles = {}  # 记录上一次接收到的角度
        self.current_angles = {"X": 0, "Y": 0, "Z": 0, "A": 0}  # 添加当前角度记录
        self.angle_offsets = {"X": 0.0, "Y": 0.0, "Z": 0.0, "A": 0.0}  # 零点偏移量
        self.raw_angles = {"X": 0.0, "Y": 0.0, "Z": 0.0, "A": 0.0}  # 原始物理角度
        self.pending_targets = {}
        self.expected_changes = {}  # 保持理论变化量
        self.realtime_deviation_history = {m: deque(maxlen=1000) for m in ["X", "Y", "Z", "A"]}
        self.expected_rotation = {"X": 0, "Y": 0, "Z": 0, "A": 0}  # 预计转动量
        self.theoretical_deviations = {m: None for m in ["X", "Y", "Z", "A"]}  # 理论偏差存储
        self.auto_calibration_enabled = False
        self.copied_step = None
        self.expected_angles = {m: 0.0 for m in ["X", "Y", "Z", "A"]}
        self.is_initializing = False
        self.active_motors = set(["X", "Y", "Z", "A"])
        self.initial_angle_base = {m: None for m in ["X", "Y", "Z", "A"]}  # 初始基准角度
        self.accumulated_rotation = {m: 0.0 for m in ["X", "Y", "Z", "A"]}  # 累积原始转动量
        self.theoretical_target = {m: None for m in ["X", "Y", "Z", "A"]}  # 理论目标角度
        self.is_first_command = True  # 是否是第一条指令
        self.running_mode = "manual"
        self.calibration_amplitude = 0.5  # 保留兼容性
        self.pid_precision = 0.5  # PID 精确控制目标阈值（度）
        self.calibration_in_progress = False  # PID 校准进行中标志
        
        # --- 新增：PID 分析器 ---
        self.pid_analyzer = PIDAnalyzer(max_history=100)
        self.pid_update_timer = QTimer()
        self.pid_update_timer.timeout.connect(self._update_pid_analysis_display)
        self.pid_update_timer.setInterval(200)  # 200ms 更新一次显示
        
        # --- 新增：关闭标志和数据包缓冲 ---
        self._closing = False  # 关闭标志，防止信号处理
        self._pending_pid_packets = []  # PID数据包缓冲
        self._chart_update_timer = QTimer()
        self._chart_update_timer.timeout.connect(self._batch_update_charts)
        self._chart_update_timer.setInterval(50)  # 20Hz 批量更新图表

        # --- 新增：加载设置 ---
        self.load_settings()

    def init_ui(self):
        self.setWindowTitle("环境现场监测系统控制程序")

        # 主布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(16)

        # 左侧导航栏
        self.nav_frame = QFrame()
        self.nav_frame.setFixedWidth(220)  # 稍微加宽以容纳新标题
        self.nav_layout = QVBoxLayout(self.nav_frame)
        self.nav_layout.setContentsMargins(0, 0, 0, 0)
        self.nav_layout.setSpacing(8)

        # ================= 控制模式 =================
        control_group = QGroupBox("泵送控制")
        control_group.setFont(QFont("Microsoft YaHei", 13))
        control_layout = QVBoxLayout(control_group)

        self.manual_btn = QPushButton("手动控制")
        self.manual_btn.setCheckable(True)
        self.manual_btn.setChecked(True)
        self.auto_btn = QPushButton("自动控制")
        self.auto_btn.setCheckable(True)

        for btn in [self.manual_btn, self.auto_btn]:
            btn.setFont(QFont("Microsoft YaHei", 13))
            btn.setFixedHeight(40)
            control_layout.addWidget(btn)

        self.nav_layout.addWidget(control_group)

        # ================= 微泵状态 =================
        self.motor_status_group = QGroupBox("泵体状态")
        self.motor_status_group.setFont(QFont("Microsoft YaHei", 13))
        motor_status_layout = QVBoxLayout(self.motor_status_group)

        self.position_btn = QPushButton("转子监控")
        self.position_btn.setCheckable(True)
        self.analysis_btn = QPushButton("运行分析")
        self.analysis_btn.setCheckable(True)

        for btn in [self.position_btn, self.analysis_btn]:
            btn.setFont(QFont("Microsoft YaHei", 13))
            btn.setFixedHeight(40)
            motor_status_layout.addWidget(btn)

        self.nav_layout.addWidget(self.motor_status_group)

        # ================= 光谱仪控制 (新增) =================
        spectro_control_group = QGroupBox("光谱仪控制")
        spectro_control_group.setFont(QFont("Microsoft YaHei", 13))
        spectro_control_layout = QVBoxLayout(spectro_control_group)

        self.spectro_btn = QPushButton("光谱分析")
        self.spectro_btn.setCheckable(True)
        self.spectro_btn.setFont(QFont("Microsoft YaHei", 13))
        self.spectro_btn.setFixedHeight(40)
        spectro_control_layout.addWidget(self.spectro_btn)
        self.nav_layout.addWidget(spectro_control_group)

        # ================= 串口设置 =================
        serial_group = QGroupBox("串口设置")
        serial_group.setFont(QFont("Microsoft YaHei", 13))
        serial_layout = QVBoxLayout(serial_group)

        # 端口选择
        self.port_combo = QComboBox()
        self.port_combo.addItems(self.get_available_ports())
        if "COM4" in self.get_available_ports():
            self.port_combo.setCurrentText("COM4")

        # 波特率选择
        self.baud_combo = QComboBox()
        self.baud_combo.addItems(['9600', '19200', '38400', '57600', '115200'])
        self.baud_combo.setCurrentText("115200")

        # 统一样式
        for widget in [self.port_combo, self.baud_combo]:
            widget.setFont(QFont("Microsoft YaHei", 16))
            widget.setFixedHeight(40)
            serial_layout.addWidget(widget)

        # 操作按钮
        self.connect_btn = QPushButton("打开串口")
        refresh_btn = QPushButton("刷新端口")

        for btn in [self.connect_btn, refresh_btn]:
            btn.setFont(QFont("Microsoft YaHei", 16))
            btn.setFixedHeight(40)
            serial_layout.addWidget(btn)

        self.nav_layout.addWidget(serial_group)
        self.nav_layout.addStretch()

        # 主内容区
        content_frame = QFrame()
        content_layout = QVBoxLayout(content_frame)
        content_layout.setContentsMargins(0, 0, 0, 0)

        # ================= 自动校准开关 (移动到电机状态组内或保持独立) =================
        self.add_auto_calibration_switch()  # 保持原位

        # ================= 标签页 =================
        self.tab_widget = QTabWidget()
        self.tab_widget.tabBar().hide()

        # 手动控制页
        self.manual_tab = QWidget()
        self.init_manual_tab()
        self.tab_widget.addTab(self.manual_tab, "")

        # 自动控制页
        self.auto_tab = QWidget()
        self.init_auto_tab()
        self.tab_widget.addTab(self.auto_tab, "")

        # 位置数据页
        self.position_tab = QWidget()
        self.init_position_tab()
        self.tab_widget.addTab(self.position_tab, "")

        # 数据分析页
        self.analysis_tab = QWidget()
        self.init_analysis_tab()
        self.tab_widget.addTab(self.analysis_tab, "")

        # 光谱分析页 (新增)
        self.spectro_tab = QWidget()
        self.init_spectro_tab()
        self.tab_widget.addTab(self.spectro_tab, "")

        content_layout.addWidget(self.tab_widget)

        # ================= 日志区域 =================
        log_group = QGroupBox("系统日志")
        log_group.setFont(QFont("Microsoft YaHei", 16))
        log_layout = QVBoxLayout(log_group)

        self.log_text = QTextEdit()
        self.log_text.setFont(QFont("Menlo", 11))
        log_layout.addWidget(self.log_text)

        clear_btn = QPushButton("清空日志")
        clear_btn.setFont(QFont("Microsoft YaHei", 16))
        clear_btn.clicked.connect(self.clear_log)
        log_layout.addWidget(clear_btn)

        content_layout.addWidget(log_group)

        # 组合布局
        main_layout.addWidget(self.nav_frame)
        main_layout.addWidget(content_frame)

        # ================= 信号连接 =================
        self.manual_btn.clicked.connect(lambda: self.switch_tab(0))
        self.auto_btn.clicked.connect(lambda: self.switch_tab(1))
        self.position_btn.clicked.connect(lambda: self.switch_tab(2))
        self.analysis_btn.clicked.connect(lambda: self.switch_tab(3))
        self.spectro_btn.clicked.connect(lambda: self.switch_tab(4))  # 新增连接

        self.connect_btn.clicked.connect(self.toggle_serial)
        refresh_btn.clicked.connect(self.refresh_serial_ports)

        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")

        # 初始刷新光谱仪设备
        if NIDAQMX_AVAILABLE:
            self._spectro_refresh_devices()

    # --- 以下是所有光谱仪相关方法 ---

    def _spectro_init_vars(self):
        """初始化所有光谱仪相关的实例变量。"""
        self.spectro_reference_voltage = None
        self.spectro_is_measuring = False
        self.spectro_task = None
        self.spectro_voltage_data = []
        self.spectro_absorbance_data = []
        self.spectro_max_data_points = 500
        self.spectro_sample_rate = 100
        self.spectro_data_log = []
        self.spectro_daq_thread = None
        self.spectro_timer = QTimer(self)
        self.spectro_timer.setTimerType(Qt.PreciseTimer)
        self.spectro_timer.timeout.connect(self._spectro_update_charts)

    def init_spectro_tab(self):
        """初始化光谱分析标签页的UI。"""
        if not NIDAQMX_AVAILABLE:
            # 如果库不可用，显示一个提示信息
            layout = QVBoxLayout(self.spectro_tab)
            label = QLabel("光谱仪功能不可用，因为缺少必要的依赖库\n(nidaqmx, pyqtgraph, scipy)。")
            label.setAlignment(Qt.AlignCenter)
            label.setFont(QFont("Microsoft YaHei", 14))
            layout.addWidget(label)
            return

        main_layout = QHBoxLayout(self.spectro_tab)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)

        # 左侧控制面板
        left_panel = QFrame()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(5, 5, 5, 5)
        left_layout.setSpacing(10)

        # 右侧图表面板
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 700])  # 设置初始大小比例

        # 填充左右面板
        self._spectro_create_device_selection_group(left_layout)
        self._spectro_create_measurement_group(left_layout)
        self._spectro_create_control_buttons(left_layout)
        left_layout.addStretch(1)

        self._spectro_create_charts_group(right_layout)

    def _spectro_create_device_selection_group(self, parent_layout):
        group = QGroupBox("设备选择")
        layout = QFormLayout()
        layout.setSpacing(8)

        self.spectro_device_combo = QComboBox()
        self.spectro_channel_combo = QComboBox()
        self.spectro_rate_spin = QSpinBox()
        self.spectro_rate_spin.setRange(1, 1000)
        self.spectro_rate_spin.setValue(self.spectro_sample_rate)
        self.spectro_rate_spin.valueChanged.connect(self._spectro_update_sample_rate)
        refresh_btn = QPushButton("刷新设备")
        refresh_btn.clicked.connect(self._spectro_refresh_devices)

        layout.addRow("设备:", self.spectro_device_combo)
        layout.addRow("通道:", self.spectro_channel_combo)
        layout.addRow("采样率 (Hz):", self.spectro_rate_spin)
        layout.addRow(refresh_btn)

        group.setLayout(layout)
        parent_layout.addWidget(group)

        self.spectro_device_combo.currentIndexChanged.connect(self._spectro_update_channels)

    def _spectro_create_measurement_group(self, parent_layout):
        group = QGroupBox("实时测量")
        layout = QFormLayout()
        layout.setSpacing(8)

        self.spectro_voltage_value = QLabel("0.0000 V")
        self.spectro_voltage_value.setStyleSheet("font-size: 20px; color: #007AFF; font-weight: bold;")
        self.spectro_absorbance_value = QLabel("0.0000")
        self.spectro_absorbance_value.setStyleSheet("font-size: 20px; color: #FF2D55; font-weight: bold;")
        self.spectro_ref_value = QLabel("未设置")
        self.spectro_ref_value.setStyleSheet("font-size: 16px; color: #5856D6;")

        layout.addRow("电压:", self.spectro_voltage_value)
        layout.addRow("吸光度:", self.spectro_absorbance_value)
        layout.addRow("参考电压:", self.spectro_ref_value)

        group.setLayout(layout)
        parent_layout.addWidget(group)

    def _spectro_create_control_buttons(self, parent_layout):
        group = QGroupBox("操作控制")
        layout = QVBoxLayout()
        layout.setSpacing(8)

        self.spectro_start_btn = QPushButton("开始测量")
        self.spectro_start_btn.clicked.connect(self._spectro_toggle_measurement)
        self.spectro_ref_btn = QPushButton("设置参考")
        self.spectro_ref_btn.clicked.connect(self._spectro_set_reference)
        self.spectro_ref_btn.setEnabled(False)
        self.spectro_clear_btn = QPushButton("清除数据")
        self.spectro_clear_btn.clicked.connect(self._spectro_clear_data)
        self.spectro_save_btn = QPushButton("保存数据")
        self.spectro_save_btn.clicked.connect(self._spectro_save_data)

        layout.addWidget(self.spectro_start_btn)
        layout.addWidget(self.spectro_ref_btn)
        layout.addWidget(self.spectro_clear_btn)
        layout.addWidget(self.spectro_save_btn)

        group.setLayout(layout)
        parent_layout.addWidget(group)

    def _spectro_create_charts_group(self, parent_layout):
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        charts_container = QWidget()
        charts_layout = QVBoxLayout(charts_container)
        charts_layout.setSpacing(15)

        # Voltage Plot
        voltage_group = QGroupBox("电压 (V)")
        v_layout = QVBoxLayout(voltage_group)
        self.spectro_voltage_plot = pg.PlotWidget()
        self.spectro_voltage_plot.setBackground('w')
        self.spectro_voltage_plot.showGrid(x=True, y=True, alpha=0.3)
        self.spectro_voltage_curve = self.spectro_voltage_plot.plot(pen=pg.mkPen(color='#007AFF', width=2))
        self.spectro_voltage_plot.setMinimumHeight(200)
        v_layout.addWidget(self.spectro_voltage_plot)
        charts_layout.addWidget(voltage_group)

        # Absorbance Plot
        absorbance_group = QGroupBox("吸光度 (Abs)")
        a_layout = QVBoxLayout(absorbance_group)
        self.spectro_absorbance_plot = pg.PlotWidget()
        self.spectro_absorbance_plot.setBackground('w')
        self.spectro_absorbance_plot.showGrid(x=True, y=True, alpha=0.3)
        self.spectro_absorbance_curve = self.spectro_absorbance_plot.plot(pen=pg.mkPen(color='#FF2D55', width=2))
        self.spectro_absorbance_plot.setMinimumHeight(200)
        a_layout.addWidget(self.spectro_absorbance_plot)
        charts_layout.addWidget(absorbance_group)

        scroll_area.setWidget(charts_container)
        parent_layout.addWidget(scroll_area)

    def _spectro_refresh_devices(self):
        try:
            system = nidaqmx.system.System.local()
            devices = system.devices
            self.spectro_device_combo.clear()
            for device in devices:
                self.spectro_device_combo.addItem(device.name)
            if self.spectro_device_combo.count() > 0:
                self.log(f"找到 {len(devices)} 个NI设备")
                self._spectro_update_channels()
            else:
                self.log("未找到NI设备")
        except Exception as e:
            self.log(f"刷新NI设备错误: {e}")

    def _spectro_update_channels(self):
        try:
            device_name = self.spectro_device_combo.currentText()
            if not device_name: return
            device = nidaqmx.system.Device(device_name)
            ai_ports = [ch.name for ch in device.ai_physical_chans]
            self.spectro_channel_combo.clear()
            self.spectro_channel_combo.addItems(ai_ports)
        except Exception as e:
            self.log(f"更新通道错误: {e}")

    def _spectro_update_sample_rate(self, rate):
        self.spectro_sample_rate = rate
        self.status_bar.showMessage(f"采样率设置为 {rate} Hz")

    def _spectro_toggle_measurement(self):
        if not self.spectro_is_measuring:
            try:
                device = self.spectro_device_combo.currentText()
                channel = self.spectro_channel_combo.currentText()
                if not device or not channel:
                    self.log("请选择设备和通道")
                    return

                if self.spectro_daq_thread and self.spectro_daq_thread.isRunning():
                    self.spectro_daq_thread.stop()

                self.spectro_daq_thread = DAQThread(device, channel, self.spectro_sample_rate)
                self.spectro_daq_thread.data_acquired.connect(self._spectro_handle_acquired_data)
                self.spectro_daq_thread.error_occurred.connect(self._spectro_handle_daq_error)
                self.spectro_daq_thread.start()

                self.spectro_is_measuring = True
                self.spectro_start_btn.setText("停止测量")
                self.spectro_ref_btn.setEnabled(True)
                self.status_bar.showMessage("测量中...")
                self.log("光谱仪开始测量")
                self.spectro_timer.start(100)  # 10Hz 更新图表
                self.spectro_start_time = time.time()
            except Exception as e:
                self.log(f"开始测量失败: {e}")
                self.spectro_is_measuring = False
                self.spectro_start_btn.setText("开始测量")
        else:
            self._spectro_stop_measurement()

    def _spectro_stop_measurement(self):
        if self.spectro_timer.isActive():
            self.spectro_timer.stop()
        if self.spectro_daq_thread and self.spectro_daq_thread.isRunning():
            self.spectro_daq_thread.stop()

        self.spectro_is_measuring = False
        self.spectro_start_btn.setText("开始测量")
        self.spectro_ref_btn.setEnabled(False)
        self.status_bar.showMessage("测量已停止")
        self.log("光谱仪停止测量")

    def _spectro_handle_acquired_data(self, voltage):
        self.spectro_voltage_data.append(voltage)
        if len(self.spectro_voltage_data) > self.spectro_max_data_points:
            self.spectro_voltage_data.pop(0)

        self.spectro_voltage_value.setText(f"{voltage:.4f} V")

        absorbance = 0.0
        if self.spectro_reference_voltage is not None and self.spectro_reference_voltage > 1e-9:
            transmittance = voltage / self.spectro_reference_voltage
            absorbance = -np.log10(transmittance) if transmittance > 0 else 0.0
            self.spectro_absorbance_value.setText(f"{absorbance:.4f}")
        else:
            self.spectro_absorbance_value.setText("N/A")

        self.spectro_absorbance_data.append(absorbance)
        if len(self.spectro_absorbance_data) > self.spectro_max_data_points:
            self.spectro_absorbance_data.pop(0)

        timestamp = time.time() - self.spectro_start_time
        self.spectro_data_log.append({
            'timestamp': timestamp,
            'voltage': voltage,
            'absorbance': absorbance,
        })

    def _spectro_handle_daq_error(self, error_msg):
        self.log(f"DAQ 错误: {error_msg}")
        self._spectro_stop_measurement()
        QMessageBox.critical(self, "测量错误", error_msg)

    def _spectro_set_reference(self):
        if self.spectro_is_measuring and len(self.spectro_voltage_data) > 0:
            avg_voltage = np.mean(self.spectro_voltage_data[-5:])
            self.spectro_reference_voltage = avg_voltage
            self.spectro_ref_value.setText(f"{self.spectro_reference_voltage:.4f} V")
            self.log(f"参考电压设置为 {self.spectro_reference_voltage:.4f} V")

    def _spectro_clear_data(self):
        self.spectro_voltage_data.clear()
        self.spectro_absorbance_data.clear()
        self.spectro_data_log.clear()
        self._spectro_update_charts()
        self.log("光谱数据已清除")

    def _spectro_update_charts(self):
        self.spectro_voltage_curve.setData(self.spectro_voltage_data)
        self.spectro_absorbance_curve.setData(self.spectro_absorbance_data)

    def _spectro_save_data(self):
        if not self.spectro_data_log:
            QMessageBox.warning(self, "保存错误", "没有数据可以保存")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"absorbance_data_{timestamp}.csv"
        path, _ = QFileDialog.getSaveFileName(self, "保存数据", filename, "CSV Files (*.csv)")
        if not path: return
        try:
            with open(path, 'w', newline='') as csvfile:
                fieldnames = ['timestamp', 'voltage', 'absorbance']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(self.spectro_data_log)
            self.log(f"光谱数据已保存至 {os.path.basename(path)}")
        except Exception as e:
            self.log(f"保存数据失败: {e}")
            QMessageBox.critical(self, "保存错误", f"文件保存失败: {e}")

    # --- 以下是原有 main.py 的方法 ---

    def add_auto_calibration_switch(self):
        """在左侧导航栏添加 PID 精确控制开关和目标阈值输入"""
        auto_cal_group = QGroupBox("PID精确控制")
        auto_cal_group.setFont(QFont("Microsoft YaHei", 13))
        auto_cal_layout = QVBoxLayout(auto_cal_group)
        # 开关布局
        switch_frame = QFrame()
        hbox = QHBoxLayout(switch_frame)
        auto_cal_label = QLabel("PID模式:")
        self.auto_cal_switch = IOSSwitch()
        hbox.addWidget(auto_cal_label)
        hbox.addWidget(self.auto_cal_switch)
        auto_cal_layout.addWidget(switch_frame)
        # 目标阈值输入框（原校准幅值）
        amplitude_frame = QFrame()
        hbox_amp = QHBoxLayout(amplitude_frame)
        amp_label = QLabel("目标阈值：")
        self.calibration_amp_input = QLineEdit()
        self.calibration_amp_input.setFixedWidth(80)
        self.calibration_amp_input.setText("0.5")
        self.calibration_amp_input.setValidator(QDoubleValidator(0.05, 2.0, 2))  # 限制输入范围0.1-5度
        # 添加单位标签
        unit_label = QLabel("°")
        hbox_amp.addWidget(amp_label)
        hbox_amp.addWidget(self.calibration_amp_input)
        hbox_amp.addWidget(unit_label)
        auto_cal_layout.addWidget(amplitude_frame)
        # 信号连接
        self.auto_cal_switch.stateChanged.connect(self.toggle_auto_calibration)
        self.calibration_amp_input.textChanged.connect(self.update_pid_precision)
        self.nav_layout.insertWidget(3, auto_cal_group)

    def update_pid_precision(self):
        """更新 PID 目标阈值（精度）"""
        try:
            value = float(self.calibration_amp_input.text())
            if 0.05 <= value <= 2.0:
                self.pid_precision = value
                self.calibration_amplitude = value  # 保持兼容性
            else:
                raise ValueError("超出范围")
        except ValueError:
            self.log("目标阈值无效，已重置为0.5°")
            self.calibration_amp_input.setText("0.5")
            self.pid_precision = 0.5
            self.calibration_amplitude = 0.5

    def toggle_auto_calibration(self, state):
        self.auto_calibration_enabled = state
        mode = "PID精确控制" if state else "传统开环"
        self.log(f"已切换至{mode}模式")

    def init_position_tab(self):
        layout = QVBoxLayout(self.position_tab)

        # 电机状态展示区 - 使用2x2网格布局
        motor_frame = QFrame()
        motor_layout = QGridLayout(motor_frame)
        motor_layout.setSpacing(10)
        self.motors = {}
        self.angle_labels = {}
        self.calibration_switches = {}  # 新增校准开关字典

        # 初始化零点标定UI字典
        self.zero_buttons = {}
        self.offset_labels = {}
        
        # 存储位置监控页的GroupBox引用，用于刷新标题
        self.position_motor_groups = {}
        
        # 2x2网格位置映射
        grid_positions = {'X': (0, 0), 'Y': (0, 1), 'Z': (1, 0), 'A': (1, 1)}
        
        for motor in ["X", "Y", "Z", "A"]:
            # 获取备注并生成标题
            note = self.settings_manager.get_pump_note(motor)
            title = f"微泵 {motor} ({note})" if note else f"微泵 {motor}"
            group = QGroupBox(title)
            self.position_motor_groups[motor] = group
            
            # 启用右键菜单
            group.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
            group.customContextMenuRequested.connect(
                lambda pos, m=motor, g=group: self.show_pump_note_menu(m, pos, g)
            )
            
            vbox = QVBoxLayout(group)
            vbox.setAlignment(Qt.AlignCenter)
            vbox.setSpacing(5)

            # 动画组件
            circle = MotorCircle()
            self.motors[motor] = circle

            # 角度显示
            label = QLabel("0.000°", alignment=Qt.AlignCenter)
            label.setStyleSheet("font-size: 24px; color: #007AFF;")
            self.angle_labels[motor] = label

            # 零点标定按钮
            zero_btn = QPushButton("设为零点")
            zero_btn.setFont(QFont("Microsoft YaHei", 10))
            zero_btn.setStyleSheet("font-size: 12px; padding: 4px;")
            zero_btn.clicked.connect(lambda checked, m=motor: self.set_current_as_zero(m))
            self.zero_buttons[motor] = zero_btn
            
            # 偏移量显示
            offset_label = QLabel("偏移: 0.0°")
            offset_label.setAlignment(Qt.AlignCenter)
            offset_label.setStyleSheet("color: #8e8e93; font-size: 11px;")
            self.offset_labels[motor] = offset_label

            # 复位开关
            switch_frame = QFrame()
            hbox = QHBoxLayout(switch_frame)
            hbox.setContentsMargins(0, 0, 0, 0)
            switch_label = QLabel("复位开关:")
            switch = IOSSwitch()
            self.calibration_switches[motor] = switch
            hbox.addWidget(switch_label)
            hbox.addWidget(switch)

            vbox.addWidget(circle, alignment=Qt.AlignCenter)
            vbox.addWidget(label)
            vbox.addWidget(zero_btn)
            vbox.addWidget(offset_label)
            vbox.addWidget(switch_frame)
            
            # 使用2x2网格布局
            row, col = grid_positions[motor]
            motor_layout.addWidget(group, row, col)

        layout.addWidget(motor_frame)

        # 控制按钮区域
        btn_frame = QFrame()
        hbox = QHBoxLayout(btn_frame)
        self.init_btn = QPushButton("微泵复位")
        self.init_btn.clicked.connect(self.start_calibration)

        # 实时角度按钮
        self.stream_btn = QPushButton("实时角度")
        self.stream_btn.setCheckable(True)
        self.stream_btn.clicked.connect(self.toggle_streaming)
        
        # 重置零点按钮
        self.reset_zero_btn = QPushButton("重置零点")
        self.reset_zero_btn.setStyleSheet("""
            QPushButton {
                font-size: 18px;
                padding: 8px;
                background-color: #ff3b30;
                color: white;
            }
            QPushButton:hover {
                background-color: #e02d24;
            }
        """)
        self.reset_zero_btn.clicked.connect(self.reset_zero_offsets)
        
        # 重置偏差按钮
        self.reset_deviation_btn = QPushButton("重置偏差")
        self.reset_deviation_btn.setStyleSheet("""
            QPushButton {
                font-size: 18px;
                padding: 8px;
                background-color: #ff9500;
                color: white;
            }
            QPushButton:hover {
                background-color: #e08600;
            }
        """)
        self.reset_deviation_btn.clicked.connect(self.reset_deviation_data)

        for btn in [self.init_btn, self.stream_btn, self.reset_zero_btn, self.reset_deviation_btn]:
            if btn not in [self.reset_zero_btn, self.reset_deviation_btn]:  # 这两个按钮已有样式
                btn.setStyleSheet("font-size: 18px; padding: 8px;")
            hbox.addWidget(btn)

        layout.addWidget(btn_frame)

    def start_calibration(self):
        """开始 PID 闭环校准流程（下位机自主完成）"""
        if not self.serial_port or not self.serial_port.is_open:
            QMessageBox.warning(self, "警告", "请先连接串口")
            return

        # 收集需要校准的电机
        motors_to_cal = ""
        for motor in ["X", "Y", "Z", "A"]:
            if self.calibration_switches[motor].isChecked():
                motors_to_cal += motor

        if not motors_to_cal:
            QMessageBox.warning(self, "提示", "请至少选择一个需要校准的电机。")
            return

        # 发送 PID 校准指令（下位机自主完成闭环控制）
        command = f"CAL{motors_to_cal}\r\n"
        if self.send_command(command):
            self.calibration_in_progress = True
            self.log(f"开始 PID 校准: {motors_to_cal}")
            self.status_bar.showMessage(f"正在校准 {motors_to_cal}...")

    def stop_calibration(self):
        """停止校准"""
        if self.send_command("CALSTOP\r\n"):
            self.calibration_in_progress = False
            self.log("校准已停止")
            self.status_bar.showMessage("校准已停止")

    def handle_calibration_message(self, data):
        """处理下位机校准状态消息"""
        if data.startswith("CAL_START"):
            self.log(f"校准开始: {data}")
            self.calibration_in_progress = True
            
        elif data.startswith("CAL_DONE:"):
            # 单个电机完成，如 CAL_DONE:X=0.32
            motor_info = data.replace("CAL_DONE:", "")
            self.log(f"电机校准完成: {motor_info}")
            
        elif data.startswith("CAL_COMPLETE"):
            # 全部完成
            self.calibration_in_progress = False
            self.log("所有电机校准完成")
            self.clear_chart()
            self.status_bar.showMessage("校准完成")
            QMessageBox.information(self, "完成", "电机 PID 校准完成")
            
        elif data.startswith("CAL_FAIL"):
            # 校准失败
            self.calibration_in_progress = False
            error_info = data.replace("CAL_FAIL:", "")
            self.log(f"校准失败: {error_info}")
            self.status_bar.showMessage("校准失败")
            QMessageBox.warning(self, "错误", f"校准失败: {error_info}")
            
        elif data.startswith("CAL_STATUS"):
            # 校准进度状态
            status_info = data.replace("CAL_STATUS:", "")
            self.log(f"校准状态: {status_info}")
            
        elif data.startswith("CAL_STOPPED"):
            self.calibration_in_progress = False
            self.log("校准已停止")
            
        elif data.startswith("CAL_ERR"):
            self.log(f"校准错误: {data}")

    def handle_pid_message(self, data):
        """处理 PID 定位模式文本消息（开始/完成/超时/失败）"""
        if data.startswith("PID_START"):
            # PID 定位开始，如 PID_START:X,target=90.0,prec=0.50
            info = data.replace("PID_START:", "")
            self.log(f"PID定位开始: {info}")
            
            try:
                parts = info.split(",")
                motor = parts[0].strip()
                target = float(parts[1].split("=")[1])
                precision = float(parts[2].split("=")[1]) if len(parts) > 2 else self.pid_precision
                
                self.pid_analyzer.start_pid_move(motor, target, precision)
                # 清空该电机的图表数据（只显示最新一次运行）
                self.pid_analysis_chart.clear_motor(motor)
                self.pid_stats_panel.update_status(motor, "运行中", "#2ca02c")
                self.pid_stats_panel.update_target(motor, target)
                
                if not self.pid_update_timer.isActive():
                    self.pid_update_timer.start()
            except Exception as e:
                self.log(f"解析 PID_START 失败: {e}")
            
        elif data.startswith("PID_DONE"):
            # PID 定位完成，如 PID_DONE:X,angle=89.80,err=0.20
            info = data.replace("PID_DONE:", "")
            self.log(f"PID定位完成: {info}")
            self.status_bar.showMessage(f"定位完成: {info}")
            
            try:
                parts = info.split(",")
                motor = parts[0].strip()
                final_angle = float(parts[1].split("=")[1])
                final_error = float(parts[2].split("=")[1])
                
                self.pid_analyzer.finish_pid_move(motor, PIDStatus.DONE, final_angle, final_error)
                self.pid_stats_panel.update_status(motor, "已完成", "#1f77b4")
                self.pid_stats_panel.update_error(motor, final_error)
                self._update_pid_history_stats()
                
                # 通知自动化线程 PID 完成
                self._notify_automation_pid_complete(motor)
                
                if not self.pid_analyzer.get_active_motors():
                    self.pid_update_timer.stop()
            except Exception as e:
                self.log(f"解析 PID_DONE 失败: {e}")
            
        elif data.startswith("PID_TIMEOUT"):
            info = data.replace("PID_TIMEOUT:", "")
            self.log(f"PID定位超时: {info}")
            self.status_bar.showMessage(f"定位超时: {info}")
            
            try:
                parts = info.split(",")
                motor = parts[0].strip()
                final_angle = float(parts[1].split("=")[1])
                final_error = float(parts[2].split("=")[1])
                
                self.pid_analyzer.finish_pid_move(motor, PIDStatus.TIMEOUT, final_angle, final_error)
                self.pid_stats_panel.update_status(motor, "超时", "#ff7f0e")
                self._update_pid_history_stats()
                
                # 通知自动化线程 PID 完成（超时也算完成）
                self._notify_automation_pid_complete(motor)
                
                if not self.pid_analyzer.get_active_motors():
                    self.pid_update_timer.stop()
            except Exception as e:
                self.log(f"解析 PID_TIMEOUT 失败: {e}")
            
        elif data.startswith("PID_FAIL"):
            info = data.replace("PID_FAIL:", "")
            self.log(f"PID定位失败: {info}")
            self.status_bar.showMessage(f"定位失败: {info}")
            
            try:
                parts = info.split("=")
                motor = parts[0].strip()
                self.pid_analyzer.finish_pid_move(motor, PIDStatus.FAILED, 0, 0)
                self.pid_stats_panel.update_status(motor, "失败", "#d62728")
                self._update_pid_history_stats()
                
                # 通知自动化线程 PID 完成（失败也算完成）
                self._notify_automation_pid_complete(motor)
                
                if not self.pid_analyzer.get_active_motors():
                    self.pid_update_timer.stop()
            except Exception as e:
                self.log(f"解析 PID_FAIL 失败: {e}")
            
        elif data.startswith("PID_STOP"):
            self.log("PID定位已停止")
            self.pid_analyzer.stop_all()
    
    def _notify_automation_pid_complete(self, motor: str) -> None:
        """
        通知自动化线程 PID 完成
        
        Args:
            motor: 完成的电机名称
        """
        try:
            if self.automation_thread and self.automation_thread.isRunning():
                self.automation_thread.notify_pid_complete(motor)
        except Exception:
            pass  # 忽略通知失败
    
    def handle_pid_packet(self, packet: dict):
        """处理 PID 二进制数据包 - 缓冲模式，由定时器批量更新图表"""
        try:
            # 检查关闭标志
            if getattr(self, '_closing', False):
                return
            
            # 检查是否已停止运行
            if not hasattr(self, 'pid_analyzer') or self.pid_analyzer is None:
                return
            
            motor = packet.get('motor', 'X')
            
            # 更新分析器（返回是否是新记录）
            is_new_record = self.pid_analyzer.update_from_packet(packet)
            
            # 如果是新记录，清空该电机的图表数据
            if is_new_record:
                if hasattr(self, 'pid_analysis_chart') and self.pid_analysis_chart is not None:
                    self.pid_analysis_chart.clear_motor(motor)
            
            # 计算相对时间
            if motor in self.pid_analyzer.active_records:
                record = self.pid_analyzer.active_records[motor]
                relative_time = time.time() - record.start_time
            else:
                relative_time = 0
            
            # 将数据包加入缓冲区，由定时器批量更新（限制缓冲区大小防止内存泄漏）
            if hasattr(self, '_pending_pid_packets'):
                # 限制缓冲区最大1000个数据包，超过则丢弃旧数据
                if len(self._pending_pid_packets) > 1000:
                    self._pending_pid_packets = self._pending_pid_packets[-500:]
                self._pending_pid_packets.append((motor, packet, relative_time))
            
            # 更新状态面板（轻量操作，可以实时更新）
            if hasattr(self, 'pid_stats_panel') and self.pid_stats_panel is not None:
                self.pid_stats_panel.update_from_packet(motor, packet)
        except RuntimeError as e:
            # Qt 对象已删除
            if "deleted" in str(e).lower() or "C++ object" in str(e):
                return
            print(f"handle_pid_packet error: {e}")
        except Exception as e:
            print(f"handle_pid_packet error: {e}")
    
    def _batch_update_charts(self):
        """批量更新图表 - 由定时器调用，降低UI刷新频率"""
        try:
            # 检查关闭标志
            if getattr(self, '_closing', False):
                return
            
            if not hasattr(self, '_pending_pid_packets') or not self._pending_pid_packets:
                return
            
            # 获取所有待处理的数据包（限制单次处理数量）
            packets = self._pending_pid_packets[:100]  # 每次最多处理100个
            del self._pending_pid_packets[:len(packets)]
            
            if not packets:
                return
            
            # 再次检查关闭标志
            if getattr(self, '_closing', False):
                return
            
            # 批量添加数据到图表（不刷新）
            if hasattr(self, 'pid_analysis_chart') and self.pid_analysis_chart is not None:
                for motor, packet, relative_time in packets:
                    if getattr(self, '_closing', False):
                        return
                    self.pid_analysis_chart.add_data_only(motor, packet, relative_time)
                
                # 一次性刷新所有曲线
                if not getattr(self, '_closing', False):
                    self.pid_analysis_chart.refresh_all_curves()
        except RuntimeError as e:
            if "deleted" in str(e).lower() or "C++ object" in str(e):
                return
        except Exception as e:
            print(f"_batch_update_charts error: {e}")

    def handle_test_result_packet(self, result: dict):
        """
        处理 PID 测试结果二进制数据包 (0xBB)
        
        Args:
            result: 包含测试结果的字典
        """
        try:
            if getattr(self, '_closing', False):
                return
            
            # 构造 TestResult 对象
            from src.core.pid_optimizer import TestResult
            test_result = TestResult(
                motor_id=result.get('motor_id', 0),
                run_index=result.get('run_index', 0),
                total_runs=result.get('total_runs', 0),
                convergence_time_ms=result.get('convergence_time_ms', 0),
                max_overshoot=result.get('max_overshoot', 0.0),
                final_error=result.get('final_error', 0.0),
                oscillation_count=result.get('oscillation_count', 0),
                smoothness_score=result.get('smoothness_score', 0),
                startup_jerk=result.get('startup_jerk', 0.0),
                total_score=result.get('total_score', 0),
            )
            
            # 传递给优化器
            if hasattr(self, 'pid_optimizer') and self.pid_optimizer:
                self.pid_optimizer.on_test_result(test_result)
            
            # 传递给单次测试
            if getattr(self, '_single_test_active', False):
                if not hasattr(self, '_single_test_results'):
                    self._single_test_results = []
                self._single_test_results.append(test_result)
            
            # 记录日志
            motor_names = ['X', 'Y', 'Z', 'A']
            motor = motor_names[result.get('motor_id', 0)]
            self.log(f"[测试结果] {motor} 轮次{result.get('run_index', 0)}: "
                     f"得分={result.get('total_score', 0)}, "
                     f"收敛={result.get('convergence_time_ms', 0)}ms")
        except Exception as e:
            print(f"handle_test_result_packet error: {e}")
    
    def handle_angle_packet(self, angles: dict):
        """
        处理角度二进制数据包 (0xCC)
        
        Args:
            angles: 包含四个电机角度的字典 {'X': float, 'Y': float, 'Z': float, 'A': float}
        """
        try:
            if getattr(self, '_closing', False):
                return
            
            # 应用零点偏移并计算相对角度
            current_angles = {}
            for motor in ['X', 'Y', 'Z', 'A']:
                raw_angle = angles.get(motor, 0.0) % 360
                # 保存原始物理角度
                self.raw_angles[motor] = raw_angle
                # 应用偏移量计算相对角度
                offset = self.angle_offsets.get(motor, 0.0)
                relative_angle = (raw_angle - offset) % 360
                current_angles[motor] = relative_angle
            
            # 计算两种偏差
            theoretical_deviations = {}
            theoretical_targets = {}
            realtime_deviations = {}
            
            for motor in ['X', 'Y', 'Z', 'A']:
                # 仅处理启用电机
                if motor not in self.active_motors:
                    theoretical_deviations[motor] = None
                    realtime_deviations[motor] = None
                    theoretical_targets[motor] = None
                    continue
                
                current = current_angles.get(motor, 0.0)
                
                # 实时偏差计算
                if motor in self.pending_targets and self.pending_targets[motor] is not None:
                    realtime_dev = (current - self.pending_targets[motor]) % 360
                    realtime_dev = realtime_dev - 360 if realtime_dev > 180 else realtime_dev
                    realtime_deviations[motor] = realtime_dev
                else:
                    realtime_deviations[motor] = None
                
                # 理论偏差计算
                if self.initial_angle_base.get(motor) is not None:
                    theoretical_target = (self.initial_angle_base[motor] +
                                          self.accumulated_rotation[motor]) % 360
                    theoretical_targets[motor] = theoretical_target
                    theoretical_dev = (current - theoretical_target) % 360
                    if theoretical_dev > 180:
                        theoretical_dev -= 360
                    theoretical_deviations[motor] = theoretical_dev
                else:
                    theoretical_deviations[motor] = None
                    theoretical_targets[motor] = None
            
            # 自动校准处理（仅自动模式）
            if self.running_mode == "auto" and self.auto_calibration_enabled:
                for motor in self.active_motors:
                    if theoretical_deviations[motor] is not None:
                        self.theoretical_deviations[motor] = theoretical_deviations[motor]
            
            # 更新界面数据
            self.current_angles.update(current_angles)
            filtered_data = {
                "current": current_angles,
                "theoretical": {k: v for k, v in theoretical_deviations.items() if k in self.active_motors},
                "realtime": {k: v for k, v in realtime_deviations.items() if k in self.active_motors},
                "targets": {k: v for k, v in theoretical_targets.items() if k in self.active_motors}
            }
            self.angle_update.emit(filtered_data)
            # 注意：图表更新已移至 update_angles 中通过节流机制处理
            # 不再在此处直接调用 chart_view.chart().update_data()
        except Exception as e:
            print(f"handle_angle_packet error: {e}")

    def format_number(self, value):
        """格式化数值，去除末尾多余的零和小数点"""
        s = "{:.3f}".format(value).rstrip('0').rstrip('.')
        return s

    def init_analysis_tab(self):
        """初始化 PID 控制分析标签页布局"""
        layout = QVBoxLayout(self.analysis_tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # ================= 使用TabWidget分隔分析和优化 =================
        self.analysis_sub_tabs = QTabWidget()
        
        # ----- 子标签1: PID实时分析 -----
        analysis_widget = QWidget()
        analysis_layout = QVBoxLayout(analysis_widget)
        analysis_layout.setContentsMargins(5, 5, 5, 5)
        
        # PID 实时状态面板
        self.pid_stats_panel = PIDStatsPanel()
        analysis_layout.addWidget(self.pid_stats_panel, stretch=0)

        # PID 分析图表
        self.pid_analysis_chart = PIDAnalysisChart()
        analysis_layout.addWidget(self.pid_analysis_chart, stretch=3)

        # 控制按钮区域
        control_frame = QFrame()
        control_layout = QHBoxLayout(control_frame)
        control_layout.setContentsMargins(5, 5, 5, 5)
        control_layout.setSpacing(10)

        button_style = """
        QPushButton {
            font-size: 12px;
            padding: 5px 10px;
            min-width: 80px;
            max-height: 30px;
            border-radius: 4px;
        }
        QPushButton:hover {
            background-color: #e0e0e0;
        }
        """

        action_buttons = [
            ("清空图表", self.clear_pid_chart),
            ("重置统计", self.reset_pid_stats),
            ("导出报告", self.export_pid_report),
            ("导出数据", self.export_pid_data),
        ]

        for btn_text, handler in action_buttons:
            btn = QPushButton(btn_text)
            btn.setStyleSheet(button_style)
            btn.clicked.connect(handler)
            control_layout.addWidget(btn)

        control_layout.addStretch(1)
        analysis_layout.addWidget(control_frame, stretch=0)

        # 历史统计面板
        history_group = QGroupBox("历史统计")
        history_group.setFont(QFont("Microsoft YaHei", 11))
        history_layout = QHBoxLayout(history_group)

        self.stats_widgets = {}
        for motor in ["X", "Y", "Z", "A"]:
            group = QGroupBox(f"微泵 {motor}")
            form = QFormLayout(group)
            form.setSpacing(4)

            labels = {
                'total_runs': QLabel("0"),
                'success_rate': QLabel("--"),
                'avg_conv_time': QLabel("--"),
                'avg_error': QLabel("--"),
            }

            for lbl in labels.values():
                lbl.setFont(QFont("Roboto Mono", 9))
                lbl.setAlignment(Qt.AlignRight)

            form.addRow("运行次数:", labels['total_runs'])
            form.addRow("成功率:", labels['success_rate'])
            form.addRow("平均收敛:", labels['avg_conv_time'])
            form.addRow("平均误差:", labels['avg_error'])

            self.stats_widgets[motor] = labels
            history_layout.addWidget(group)

        analysis_layout.addWidget(history_group, stretch=0)
        
        self.analysis_sub_tabs.addTab(analysis_widget, "PID实时分析")
        
        # ----- 子标签2: PID参数优化 -----
        self.pid_optimizer_panel = PIDOptimizerPanel()
        self._init_pid_optimizer()
        self.analysis_sub_tabs.addTab(self.pid_optimizer_panel, "PID参数优化")
        
        layout.addWidget(self.analysis_sub_tabs)
        
        # ================= 兼容旧版图表（隐藏） =================
        self.chart_view = QChartView(AnalysisChart())
        self.chart_view.setVisible(False)
        layout.addWidget(self.chart_view)
    
    def _init_pid_optimizer(self):
        """初始化PID优化器"""
        self.pid_optimizer = PatternSearchOptimizer()
        
        # 设置串口发送回调
        self.pid_optimizer.set_send_callback(self.send_command)
        
        # 连接优化器信号
        self.pid_optimizer.progress_updated.connect(self.pid_optimizer_panel.update_progress)
        self.pid_optimizer.score_updated.connect(self.pid_optimizer_panel.update_score)
        self.pid_optimizer.state_changed.connect(self.pid_optimizer_panel.on_state_changed)
        self.pid_optimizer.optimization_finished.connect(self._on_optimization_finished)
        self.pid_optimizer.error_occurred.connect(lambda msg: self.log(f"优化器错误: {msg}"))
        
        # 连接面板信号
        self.pid_optimizer_panel.start_optimization.connect(self._start_pid_optimization)
        self.pid_optimizer_panel.stop_optimization.connect(self._stop_pid_optimization)
        self.pid_optimizer_panel.pause_optimization.connect(self.pid_optimizer.pause)
        self.pid_optimizer_panel.resume_optimization.connect(self.pid_optimizer.resume)
        self.pid_optimizer_panel.apply_params.connect(self._apply_pid_params)
        self.pid_optimizer_panel.single_test.connect(self._run_single_pid_test)
        self.pid_optimizer_panel.export_data.connect(self._export_pid_optimization_data)
        
        # 初始化单次测试状态
        self._single_test_active = False
        self._single_test_results = []
        self._single_test_params = {}
    
    def _start_pid_optimization(self, config: dict):
        """开始PID优化"""
        if not self.serial_port or not self.serial_port.is_open:
            QMessageBox.warning(self, "警告", "请先打开串口连接")
            self.pid_optimizer_panel.on_state_changed('idle')
            return
        
        # 配置优化器
        self.pid_optimizer.configure(
            test_motor=config.get('test_motor', 'X'),
            test_angle=config.get('test_angle', 45.0),
            test_runs=config.get('test_runs', 5),
            max_iterations=config.get('max_iterations', 50),
            initial_step=config.get('initial_step', 0.02),
            min_step=config.get('min_step', 0.005)
        )
        
        # 创建初始参数
        initial = config.get('initial_params', {})
        initial_params = PIDParams(
            Kp=initial.get('Kp', 0.14),
            Ki=initial.get('Ki', 0.015),
            Kd=initial.get('Kd', 0.06)
        )
        
        self.log(f"开始PID参数优化: 电机={config.get('test_motor')}, 角度={config.get('test_angle')}°")
        self.pid_optimizer.start(initial_params)
    
    def _stop_pid_optimization(self):
        """停止PID优化和单次测试"""
        # 停止优化器
        self.pid_optimizer.stop()
        
        # 停止单次测试
        if getattr(self, '_single_test_active', False):
            self._single_test_active = False
            self.send_command("PIDTESTSTOP\r\n")
            self.pid_optimizer_panel.on_single_test_complete()
        
        self.log("PID参数优化/测试已停止")
    
    def _on_optimization_finished(self, result: dict):
        """优化完成处理"""
        self.pid_optimizer_panel.on_optimization_finished(result)
        
        best = result.get('best_params', {})
        self.log(f"PID优化完成! 最优参数: Kp={best.get('Kp', 0):.4f}, Ki={best.get('Ki', 0):.5f}, Kd={best.get('Kd', 0):.4f}")
        self.log(f"最优得分: {result.get('best_score', 0):.1f}, 迭代次数: {result.get('iterations', 0)}")
        
        # 更新历史记录
        for record in self.pid_optimizer.get_history_summary():
            self.pid_optimizer_panel.add_history_record(record)
    
    def _apply_pid_params(self, params: dict):
        """应用PID参数到下位机"""
        if not self.serial_port or not self.serial_port.is_open:
            QMessageBox.warning(self, "警告", "请先打开串口连接")
            return
        
        cmd = f"PIDCFG:{params.get('Kp', 0.14):.4f},{params.get('Ki', 0.015):.5f},{params.get('Kd', 0.06):.4f}\r\n"
        if self.send_command(cmd):
            self.log(f"已应用PID参数: Kp={params.get('Kp'):.4f}, Ki={params.get('Ki'):.5f}, Kd={params.get('Kd'):.4f}")
    
    def _run_single_pid_test(self, config: dict):
        """执行单次PID测试（用于调试）"""
        if not self.serial_port or not self.serial_port.is_open:
            QMessageBox.warning(self, "警告", "请先打开串口连接")
            self.pid_optimizer_panel.on_single_test_complete()
            return
        
        params = config.get('params', {})
        motor = config.get('test_motor', 'X')
        angle = config.get('test_angle', 45.0)
        runs = config.get('test_runs', 1)
        
        # 保存单次测试参数，用于结果回调
        self._single_test_params = params.copy()
        self._single_test_runs = runs
        self._single_test_results = []
        self._single_test_active = True
        
        # 先配置PID参数
        cfg_cmd = f"PIDCFG:{params.get('Kp', 0.14):.4f},{params.get('Ki', 0.015):.5f},{params.get('Kd', 0.06):.4f}\r\n"
        self.log(f"[单次测试] 发送PID配置: {cfg_cmd.strip()}")
        if not self.send_command(cfg_cmd):
            self.log("[单次测试] PID配置发送失败")
            self._single_test_active = False
            self.pid_optimizer_panel.on_single_test_complete()
            return
        
        # 等待200ms后发送测试指令
        def send_test():
            test_cmd = f"PIDTEST:{motor},{angle:.1f},{runs}\r\n"
            self.log(f"[单次测试] 发送测试指令: {test_cmd.strip()}")
            if self.send_command(test_cmd):
                self.log(f"[单次测试] 测试已启动: 电机={motor}, 角度={angle}°, 次数={runs}")
            else:
                self.log("[单次测试] 测试指令发送失败")
                self._single_test_active = False
                self.pid_optimizer_panel.on_single_test_complete()
        
        QTimer.singleShot(200, send_test)
    
    def _on_single_test_complete(self):
        """单次测试完成处理"""
        results = getattr(self, '_single_test_results', [])
        params = getattr(self, '_single_test_params', {})
        
        if results:
            # 计算平均得分
            avg_score = sum(r.total_score for r in results) / len(results)
            
            # 更新面板
            result_data = {
                'score': avg_score,
                'Kp': params.get('Kp', 0.14),
                'Ki': params.get('Ki', 0.015),
                'Kd': params.get('Kd', 0.06),
                'runs': len(results)
            }
            self.pid_optimizer_panel.on_single_test_result(result_data)
            self.log(f"[单次测试] 完成! 平均得分: {avg_score:.1f}, 测试次数: {len(results)}")
        else:
            self.log("[单次测试] 完成，但未收到测试结果")
        
        # 通知面板测试完成
        self.pid_optimizer_panel.on_single_test_complete()
        
        # 清理状态
        self._single_test_results = []
        self._single_test_params = {}

    def _export_pid_optimization_data(self):
        """导出PID优化数据到CSV文件"""
        from PySide6.QtWidgets import QFileDialog
        import csv
        from datetime import datetime
        
        # 获取优化历史
        history = self.pid_optimizer.get_history_summary()
        if not history:
            QMessageBox.information(self, "提示", "没有可导出的优化数据")
            return
        
        # 选择保存路径
        default_name = f"pid_optimization_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出PID优化数据", default_name, "CSV文件 (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                # 写入表头
                writer.writerow(['迭代', 'Kp', 'Ki', 'Kd', '原始得分', '调整得分', '最大过冲°', '收敛时间ms', 'RSD%'])
                # 写入数据
                for record in history:
                    writer.writerow([
                        record.get('index', 0),
                        f"{record.get('Kp', 0):.4f}",
                        f"{record.get('Ki', 0):.5f}",
                        f"{record.get('Kd', 0):.4f}",
                        f"{record.get('avg_score', 0):.1f}",
                        f"{record.get('adjusted_score', record.get('avg_score', 0)):.1f}",
                        f"{record.get('max_overshoot', 0):.2f}",
                        f"{record.get('avg_conv_time', 0):.0f}",
                        f"{record.get('convergence_rsd', 0):.1f}"
                    ])
                
                # 写入最优参数
                writer.writerow([])
                writer.writerow(['最优参数（贝叶斯优化 + 非线性惩罚）'])
                if self.pid_optimizer.best_params:
                    writer.writerow(['Kp', f"{self.pid_optimizer.best_params.Kp:.4f}"])
                    writer.writerow(['Ki', f"{self.pid_optimizer.best_params.Ki:.5f}"])
                    writer.writerow(['Kd', f"{self.pid_optimizer.best_params.Kd:.4f}"])
                    writer.writerow(['最优得分（惩罚后）', f"{self.pid_optimizer.best_score:.1f}"])
            
            self.log(f"[导出] PID优化数据已导出到: {file_path}")
            QMessageBox.information(self, "成功", f"数据已导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {e}")

    def clear_chart(self):
        # 清空图表数据
        # 清空旧版图表
        self.chart_view.chart().clear()
        # 重置偏差数据集
        self.deviation_data = {m: deque(maxlen=100) for m in ["X", "Y", "Z", "A"]}
        self.theoretical_deviations = {m: None for m in ["X", "Y", "Z", "A"]}
        self.target_angles = {m: None for m in ["X", "Y", "Z", "A"]}
        self.expected_rotation = {m: 0.0 for m in ["X", "Y", "Z", "A"]}
        # 重置统计数据和颜色
        for motor in ["X", "Y", "Z", "A"]:
            labels = self.stats_widgets[motor]
            labels['total_runs'].setText("0")
            labels['success_rate'].setText("--")
            labels['avg_conv_time'].setText("--")
            labels['avg_error'].setText("--")
            for label in labels.values():
                label.setStyleSheet("color: black;")
        self.log("图表和统计数据已重置")
    
    def clear_pid_chart(self):
        """清空 PID 分析图表"""
        self.pid_analysis_chart.clear_all()
        self.pid_analyzer.clear_realtime_data()
        self.pid_stats_panel.reset_all()
        self.log("PID 分析图表已清空")
    
    def reset_pid_stats(self):
        """重置 PID 统计数据"""
        reply = QMessageBox.question(
            self, "确认重置",
            "确定要重置所有 PID 历史统计数据吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.pid_analyzer.reset()
            self.clear_pid_chart()
            self._update_pid_history_stats()
            self.log("PID 统计数据已重置")
    
    def export_pid_report(self):
        """导出 PID 分析报告"""
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装 pandas 库：pip install pandas")
            return
        
        # 收集统计数据
        report_data = []
        for motor in ["X", "Y", "Z", "A"]:
            stats = self.pid_analyzer.get_stats_summary(motor)
            stats['motor'] = motor
            report_data.append(stats)
        
        if not any(d['total_runs'] > 0 for d in report_data):
            QMessageBox.warning(self, "警告", "没有可导出的 PID 运行数据")
            return
        
        # 保存对话框
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pid_report_{timestamp}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 PID 报告", filename,
            "Excel文件 (*.xlsx)"
        )
        
        if path:
            try:
                df = pd.DataFrame(report_data)
                df = df[['motor', 'total_runs', 'success_rate', 'avg_convergence_time', 
                         'min_convergence_time', 'max_convergence_time', 
                         'avg_final_error', 'max_final_error', 'timeout_count', 'fail_count']]
                df.columns = ['电机', '运行次数', '成功率', '平均收敛时间', 
                              '最短收敛时间', '最长收敛时间',
                              '平均最终误差', '最大最终误差', '超时次数', '失败次数']
                df.to_excel(path, index=False)
                self.log(f"PID 报告已导出至 {path}")
            except Exception as e:
                QMessageBox.critical(self, "导出错误", f"文件写入失败: {str(e)}")
    
    def export_pid_data(self):
        """导出 PID 控制数据到 Excel 文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装 openpyxl 库：pip install openpyxl")
            return
        
        # 保存对话框
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pid_data_{timestamp}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 PID 数据", filename, "Excel 文件 (*.xlsx)"
        )
        
        if not path:
            return
        
        try:
            wb = Workbook()
            
            # 样式定义
            header_font = Font(name='Arial', size=11, bold=True)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='Arial', size=10)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            def apply_header_style(ws, row=1):
                for cell in ws[row]:
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
            
            def apply_data_style(ws, start_row=2):
                for row in ws.iter_rows(min_row=start_row):
                    for cell in row:
                        cell.font = data_font
                        cell.alignment = center_align
                        cell.border = thin_border
            
            def auto_column_width(ws, min_width=10, max_width=25):
                for column_cells in ws.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max(max_length + 2, min_width), max_width)
                    ws.column_dimensions[column].width = adjusted_width
            
            # ========== Sheet 1: Summary (统计摘要) ==========
            ws_summary = wb.active
            ws_summary.title = "Summary"
            headers = ["Motor", "Total Runs", "Success Rate", "Avg Conv Time (s)", 
                      "Min Conv Time (s)", "Max Conv Time (s)", "Avg Error (°)", 
                      "Max Error (°)", "Timeout", "Failed"]
            ws_summary.append(headers)
            
            for motor in ["X", "Y", "Z", "A"]:
                stats = self.pid_analyzer.stats[motor]
                row = [
                    motor,
                    stats.total_runs,
                    f"{stats.success_rate:.1f}%",
                    f"{stats.avg_convergence_time:.3f}" if stats.successful_runs > 0 else "-",
                    f"{stats.min_convergence_time:.3f}" if stats.min_convergence_time else "-",
                    f"{stats.max_convergence_time:.3f}" if stats.max_convergence_time else "-",
                    f"{stats.avg_final_error:.3f}" if stats.successful_runs > 0 else "-",
                    f"{stats.max_final_error:.3f}" if stats.max_final_error > 0 else "-",
                    stats.timeout_runs,
                    stats.failed_runs
                ]
                ws_summary.append(row)
            
            apply_header_style(ws_summary)
            apply_data_style(ws_summary)
            auto_column_width(ws_summary)
            ws_summary.freeze_panes = 'A2'
            
            # ========== Sheet 2: Error Distribution (误差分布) ==========
            ws_error_dist = wb.create_sheet("Error Distribution")
            ws_error_dist.append(["Index", "Motor", "Final Error (°)"])
            
            idx = 1
            for motor in ["X", "Y", "Z", "A"]:
                for error in self.pid_analyzer.stats[motor].error_distribution:
                    ws_error_dist.append([idx, motor, f"{error:.4f}"])
                    idx += 1
            
            apply_header_style(ws_error_dist)
            apply_data_style(ws_error_dist)
            auto_column_width(ws_error_dist)
            ws_error_dist.freeze_panes = 'A2'
            
            # ========== Sheet 3: Run History (运行历史) ==========
            ws_history = wb.create_sheet("Run History")
            ws_history.append(["Run ID", "Motor", "Target (°)", "Precision (°)", 
                              "Duration (s)", "Final Angle (°)", "Final Error (°)", "Status"])
            
            run_id = 1
            for motor in ["X", "Y", "Z", "A"]:
                for record in self.pid_analyzer.history[motor]:
                    row = [
                        run_id,
                        motor,
                        f"{record.target_angle:.2f}",
                        f"{record.precision:.2f}",
                        f"{record.duration:.3f}" if record.duration else "-",
                        f"{record.final_angle:.3f}" if record.final_angle is not None else "-",
                        f"{record.final_error:.3f}" if record.final_error is not None else "-",
                        record.status.value
                    ]
                    ws_history.append(row)
                    run_id += 1
            
            apply_header_style(ws_history)
            apply_data_style(ws_history)
            auto_column_width(ws_history)
            ws_history.freeze_panes = 'A2'
            
            # ========== Sheet 4: Realtime Position (实时位置追踪，最多20000点) ==========
            ws_position = wb.create_sheet("Realtime Position")
            pos_headers = ["Time (s)"]
            for motor in ["X", "Y", "Z", "A"]:
                pos_headers.extend([f"{motor}_Target", f"{motor}_Actual", f"{motor}_Theo"])
            ws_position.append(pos_headers)
            
            # 使用pid_analyzer的导出数据（大缓冲区）
            all_times = set()
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_position_data(motor):
                    all_times.add(round(data[0], 3))
            
            # 构建数据字典
            pos_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_position_data(motor):
                    t = round(data[0], 3)
                    pos_dict[motor][t] = (data[1], data[2], data[3])  # target, actual, theo
            
            for t in sorted(all_times):
                row = [f"{t:.3f}"]
                for motor in ["X", "Y", "Z", "A"]:
                    if t in pos_dict[motor]:
                        target, actual, theo = pos_dict[motor][t]
                        row.extend([f"{target:.3f}", f"{actual:.3f}", f"{theo:.3f}"])
                    else:
                        row.extend(["-", "-", "-"])
                ws_position.append(row)
            
            apply_header_style(ws_position)
            apply_data_style(ws_position)
            auto_column_width(ws_position)
            ws_position.freeze_panes = 'A2'
            
            # ========== Sheet 5: Realtime PID Output (最多20000点) ==========
            ws_output = wb.create_sheet("Realtime Output")
            ws_output.append(["Time (s)", "X_Output (RPM)", "Y_Output (RPM)", 
                             "Z_Output (RPM)", "A_Output (RPM)"])
            
            all_times = set()
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_output_data(motor):
                    all_times.add(round(data[0], 3))
            
            out_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_output_data(motor):
                    out_dict[motor][round(data[0], 3)] = data[1]
            
            for t in sorted(all_times):
                row = [f"{t:.3f}"]
                for motor in ["X", "Y", "Z", "A"]:
                    if t in out_dict[motor]:
                        row.append(f"{out_dict[motor][t]:.3f}")
                    else:
                        row.append("-")
                ws_output.append(row)
            
            apply_header_style(ws_output)
            apply_data_style(ws_output)
            auto_column_width(ws_output)
            ws_output.freeze_panes = 'A2'
            
            # ========== Sheet 6: Realtime Error (最多20000点) ==========
            ws_error = wb.create_sheet("Realtime Error")
            ws_error.append(["Time (s)", "X_Error (°)", "Y_Error (°)", 
                            "Z_Error (°)", "A_Error (°)"])
            
            # 使用pid_analyzer的导出数据（大缓冲区）
            all_times = set()
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_error_data(motor):
                    all_times.add(round(data[0], 3))
            
            err_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_error_data(motor):
                    err_dict[motor][round(data[0], 3)] = data[1]
            
            for t in sorted(all_times):
                row = [f"{t:.3f}"]
                for motor in ["X", "Y", "Z", "A"]:
                    if t in err_dict[motor]:
                        row.append(f"{err_dict[motor][t]:.3f}")
                    else:
                        row.append("-")
                ws_error.append(row)
            
            apply_header_style(ws_error)
            apply_data_style(ws_error)
            auto_column_width(ws_error)
            ws_error.freeze_panes = 'A2'
            
            # ========== Sheet 7: Realtime Load (最多20000点) ==========
            ws_load = wb.create_sheet("Realtime Load")
            ws_load.append(["Time (s)", "X_Load (°)", "Y_Load (°)", 
                           "Z_Load (°)", "A_Load (°)"])
            
            # 使用pid_analyzer的导出数据（大缓冲区）
            all_times = set()
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_load_data(motor):
                    all_times.add(round(data[0], 3))
            
            load_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
            for motor in ["X", "Y", "Z", "A"]:
                for data in self.pid_analyzer.get_export_load_data(motor):
                    load_dict[motor][round(data[0], 3)] = data[1]
            
            for t in sorted(all_times):
                row = [f"{t:.3f}"]
                for motor in ["X", "Y", "Z", "A"]:
                    if t in load_dict[motor]:
                        row.append(f"{load_dict[motor][t]:.3f}")
                    else:
                        row.append("-")
                ws_load.append(row)
            
            apply_header_style(ws_load)
            apply_data_style(ws_load)
            auto_column_width(ws_load)
            ws_load.freeze_panes = 'A2'
            
            # 保存文件
            wb.save(path)
            self.log(f"PID 数据已导出至 {path}")
            
        except Exception as e:
            QMessageBox.critical(self, "导出错误", f"数据导出失败: {str(e)}")
    
    def _update_pid_analysis_display(self):
        """更新 PID 分析显示（定时器回调）"""
        # 更新活动电机的实时状态
        for motor in ["X", "Y", "Z", "A"]:
            status = self.pid_analyzer.get_motor_status(motor)
            
            if status == PIDStatus.RUNNING:
                self.pid_stats_panel.update_status(motor, "运行中", "#2ca02c")
                # 获取最新误差数据
                error_data = self.pid_analyzer.get_realtime_error_data(motor)
                if error_data:
                    latest_error = error_data[-1][1]
                    self.pid_stats_panel.update_error(motor, latest_error)
            elif status == PIDStatus.DONE:
                self.pid_stats_panel.update_status(motor, "已完成", "#1f77b4")
            elif status == PIDStatus.TIMEOUT:
                self.pid_stats_panel.update_status(motor, "超时", "#ff7f0e")
            elif status == PIDStatus.FAILED:
                self.pid_stats_panel.update_status(motor, "失败", "#d62728")
            else:
                self.pid_stats_panel.update_status(motor, "空闲", "#888888")
    
    def _update_pid_history_stats(self):
        """更新 PID 历史统计显示"""
        for motor in ["X", "Y", "Z", "A"]:
            stats = self.pid_analyzer.stats[motor]
            labels = self.stats_widgets[motor]
            
            labels['total_runs'].setText(str(stats.total_runs))
            labels['success_rate'].setText(f"{stats.success_rate:.1f}%")
            labels['avg_conv_time'].setText(f"{stats.avg_convergence_time:.2f}s")
            labels['avg_error'].setText(f"{stats.avg_final_error:.2f}°")

    def save_chart_image(self):
        """保存为图片文件"""
        options = QFileDialog.Options()
        path, _ = QFileDialog.getSaveFileName(
            self, "保存图表", "",
            "PNG图像 (*.png);;JPEG图像 (*.jpg)",
            options=options
        )

        if path:
            pixmap = self.chart_view.grab()
            if pixmap.save(path):
                self.log(f"图表已保存至 {path}")
            else:
                QMessageBox.warning(self, "错误", "图片保存失败")

    def export_chart_data(self):
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装pandas库：pip install pandas")
            return
        # 获取选择的电机
        selected = self.export_motor_combo.currentText()
        motor_map = {
            "全部": ["X", "Y", "Z", "A"],
            "X轴": ["X"],
            "Y轴": ["Y"],
            "Z轴": ["Z"],
            "A轴": ["A"]
        }
        selected_motors = motor_map[selected]

        # 获取数据并对齐长度
        chart_data = self.chart_view.chart().get_chart_data()
        max_length = max(len(chart_data[m]) for m in selected_motors if chart_data[m]) if any(
            chart_data.values()) else 0
        if max_length == 0:
            QMessageBox.warning(self, "警告", "当前没有可导出的数据")
            return

        # 填充缺失值为NaN
        data_dict = {}
        for motor in selected_motors:
            values = chart_data[motor]
            if len(values) < max_length:
                values += [float('nan')] * (max_length - len(values))
            data_dict[f"{motor}轴偏差"] = values

        # 保存对话框
        options = QFileDialog.Options()
        path, _ = QFileDialog.getSaveFileName(
            self, "导出数据", "",
            "Excel文件 (*.xlsx)",
            options=options
        )

        if path:
            try:
                df = pd.DataFrame(data_dict)
                df.index.name = "采样序列"
                with pd.ExcelWriter(path) as writer:
                    df.to_excel(writer)
                self.log(f"数据已导出至 {path}")
            except Exception as e:
                QMessageBox.critical(self, "导出错误", f"文件写入失败: {str(e)}")

    def import_chart_data(self):
        """支持导入单轴Excel数据"""
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装pandas库：pip install pandas")
            return
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", "",
            "Excel文件 (*.xlsx);;CSV文件 (*.csv)",
            options=options
        )
        if not file_path:
            return
        try:
            # 读取数据文件
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path, index_col=0)
            else:
                df = pd.read_csv(file_path, index_col=0)
            # 识别有效数据列
            valid_columns = []
            motor_mapping = {
                'X轴偏差': 'X',
                'Y轴偏差': 'Y',
                'Z轴偏差': 'Z',
                'A轴偏差': 'A'
            }

            # 检查至少包含一个有效列
            for col in df.columns:
                if col in motor_mapping:
                    valid_columns.append(col)

            if not valid_columns:
                raise ValueError("未检测到有效偏差数据列（列名示例：X轴偏差）")
            # 弹窗选择导入模式
            mode = self._show_import_dialog()
            if not mode:
                return
            # 处理数据导入
            for col in valid_columns:
                motor = motor_mapping[col]
                data = df[col].dropna().tolist()

                if mode == "replace":
                    # 替换模式：清空现有数据
                    self.chart_view.chart().data[motor].clear()
                    self.chart_view.chart().data[motor].extend(data)
                else:
                    # 追加模式：添加到现有数据尾部
                    self.chart_view.chart().data[motor].extend(data)

                # 更新数据点（保留最近200个）
                points = [QPointF(x, y) for x, y in enumerate(self.chart_view.chart().data[motor])]
                self.chart_view.chart().series[motor].replace(points[-200:])
            # 刷新显示
            self.chart_view.chart().auto_scale_axes()
            self._update_stats_after_import(df, valid_columns)
            self.log(f"成功导入{len(valid_columns)}轴数据")
        except Exception as e:
            QMessageBox.critical(self, "导入错误", f"数据导入失败: {str(e)}")

    def _show_import_dialog(self):
        """显示导入选项对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("导入选项")
        layout = QVBoxLayout(dialog)

        # 模式选择
        mode_group = QButtonGroup(dialog)
        rb_replace = QRadioButton("替换现有数据", dialog)
        rb_append = QRadioButton("追加到现有数据", dialog)
        rb_replace.setChecked(True)

        mode_group.addButton(rb_replace)
        mode_group.addButton(rb_append)

        # 确认按钮
        btn_confirm = QPushButton("确认导入", dialog)
        btn_confirm.clicked.connect(dialog.accept)

        layout.addWidget(QLabel("请选择数据导入模式:"))
        layout.addWidget(rb_replace)
        layout.addWidget(rb_append)
        layout.addWidget(btn_confirm)

        if dialog.exec() == QDialog.Accepted:
            return "replace" if rb_replace.isChecked() else "append"
        return None

    def _update_stats_after_import(self, df, valid_columns):
        """更新统计信息（仅更新导入的轴）"""
        for col in valid_columns:
            motor = col[0]  # 从"X轴偏差"提取X
            data = df[col].dropna()

            if not data.empty:
                labels = self.stats_widgets[motor]
                labels['current'].setText(f"{data.iloc[-1]:.2f}°")
                labels['average'].setText(f"{data.mean():.2f}°")
                labels['dev_rate'].setText(
                    f"{(data.iloc[-1] / data.mean() * 100 if data.mean() != 0 else 0):.1f}%"
                )

    def _update_all_stats(self, df):
        """更新所有统计信息"""
        self.cumulative_deviations = {m: 0.0 for m in ["X", "Y", "Z", "A"]}

        for motor in ["X", "Y", "Z", "A"]:
            col = f"{motor}轴偏差"
            if col not in df.columns: continue
            data = df[col].dropna()

            if not data.empty:
                # 计算统计指标
                current = data.iloc[-1]
                avg = data.mean()

                # 更新显示
                labels = self.stats_widgets[motor]
                labels['current'].setText(f"{current:.2f}°")
                labels['average'].setText(f"{avg:.2f}°")
                labels['dev_rate'].setText(f"{(current / avg * 100 if avg != 0 else 0):.1f}%")

    def toggle_streaming(self, checked: bool):
        """
        切换实时角度流模式
        
        Args:
            checked: 是否开启实时流
        """
        if checked:
            self.send_command("ANGLESTREAM_START\r\n")
            self.stream_btn.setText("停止实时")
            self.stream_btn.setStyleSheet(
                "font-size: 18px; padding: 8px; background-color: #4CAF50; color: white;"
            )
        else:
            self.send_command("ANGLESTREAM_STOP\r\n")
            self.stream_btn.setText("实时角度")
            self.stream_btn.setStyleSheet("font-size: 18px; padding: 8px;")

    def switch_tab(self, index):
        self.tab_widget.setCurrentIndex(index)
        # 更新导航按钮的选中状态
        buttons = [self.manual_btn, self.auto_btn, self.position_btn, self.analysis_btn, self.spectro_btn]
        for i, btn in enumerate(buttons):
            btn.setChecked(i == index)

    def update_angles(self, data):
        """更新角度显示（适配新版数据结构）- 带节流机制"""
        try:
            if getattr(self, '_closing', False):
                return
            
            # 更新所有电机的圆形动画和角度标签
            for motor in ["X", "Y", "Z", "A"]:
                current_angle = data["current"].get(motor, 0)
                self.motors[motor].set_angle(current_angle % 360)
                self.angle_labels[motor].setText(f"{current_angle:.3f}°")
            
            # 节流：每100ms更新一次图表（10Hz）
            current_time = time.time()
            if not hasattr(self, '_last_chart_update_time'):
                self._last_chart_update_time = 0
            
            if current_time - self._last_chart_update_time >= 0.1:
                self._last_chart_update_time = current_time
                try:
                    self.chart_view.chart().update_data(data)
                except Exception:
                    pass
        except Exception as e:
            print(f"角度更新异常: {str(e)}")

    def _create_centered_item(self, text):
        """创建居中显示的表格项"""
        item = QTableWidgetItem(str(text))
        item.setTextAlignment(Qt.AlignCenter)
        return item

    def handle_serial_data(self, data: str):
        """
        处理串口文本数据
        
        Args:
            data: 接收到的文本数据行
        """
        # PID测试相关消息处理
        if data.startswith("PIDTEST_"):
            self._handle_pid_test_message(data)
            return
        
        # PID配置确认
        if data.startswith("PIDCFG_"):
            self.log(f"PID配置: {data}")
            return
        
        # PID参数查询响应
        if data.startswith("PIDPARAM:"):
            self.log(f"当前PID参数: {data}")
            return
        
        # 角度流控制响应
        if data.startswith("ANGLESTREAM_"):
            self.log(f"角度流: {data}")
            return
        
        # 校准相关消息
        if data.startswith("CAL"):
            self.handle_calibration_message(data)
            return
        
        # PID定位模式消息
        if data.startswith("PID_"):
            self.handle_pid_message(data)
            return
        
        # 下位机忙碌状态
        if data.startswith("BUSY"):
            self.log(f"下位机: {data}")
            return
        
        # 流模式状态（兼容旧版）
        if data.startswith("STREAM"):
            self.log(f"流模式: {data}")
            return
        
        # 电机状态消息（调试用）
        if data.startswith("Motor"):
            self.log(f"电机: {data}")
            return
        
        # 其他消息记录
        self.log(f"接收: {data}")
    
    def _handle_pid_test_message(self, data: str):
        """
        处理PID测试相关的文本消息
        
        Args:
            data: PID测试消息
        """
        if data.startswith("PIDTEST_START:"):
            # 测试开始
            self.log(f"[PID测试] 开始: {data}")
        
        elif data.startswith("PIDTEST_RUN:"):
            # 测试轮次
            run_index = data.split(':')[1] if ':' in data else '?'
            self.log(f"[PID测试] 执行轮次 {run_index}")
        
        elif data.startswith("PIDTEST_RESULT:"):
            # 文本格式测试结果（作为备份，主要使用二进制包）
            self.log(f"[PID测试] 结果: {data}")
        
        elif data.startswith("PIDTEST_DONE:"):
            # 测试完成
            motor = data.split(':')[1] if ':' in data else '?'
            self.log(f"[PID测试] 完成: 电机 {motor}")
            
            # 通知优化器测试完成
            if hasattr(self, 'pid_optimizer') and self.pid_optimizer:
                self.pid_optimizer.on_test_done()
            
            # 单次测试完成
            if getattr(self, '_single_test_active', False):
                self._single_test_active = False
                self._on_single_test_complete()
        
        elif data.startswith("PIDTEST_STOPPED"):
            # 测试被停止
            self.log("[PID测试] 已停止")
            
            # 如果单次测试被停止
            if getattr(self, '_single_test_active', False):
                self._single_test_active = False
                self.pid_optimizer_panel.on_single_test_complete()
        
        elif data.startswith("PIDTEST_ERR:"):
            # 测试错误
            error = data.split(':')[1] if ':' in data else '未知错误'
            self.log(f"[PID测试] 错误: {error}")

    def update_stats_panel(self, data):
        """更新统计面板（兼容旧版调用）- 已优化，不再直接更新图表"""
        # 图表更新已移至节流机制，此处不再调用
        pass

    def _set_stat_na(self, motor):
        """设置未启用电机的统计显示"""
        labels = self.stats_widgets[motor]
        for key in labels.keys():
            labels[key].setText("--")
            labels[key].setStyleSheet("color: gray;")

    def _update_stat_color(self, labels, current_dev):
        """根据偏差值更新统计颜色"""
        dev_abs = abs(current_dev)
        color = "black"  # 默认黑色

        if dev_abs > 5.0:  # 严重偏差
            color = "#FF0000"  # 红色
        elif dev_abs > 2.0:  # 警告偏差
            color = "#FFA500"  # 橙色

        # 只更新数值标签的颜色
        for label_key in labels.keys():
            labels[label_key].setStyleSheet(f"color: {color};")

    def set_size_policy(self, h_policy, v_policy):
        """通用尺寸策略设置方法"""
        for widget in self.findChildren(QWidget):
            if isinstance(widget, (QGroupBox, QFrame)):
                widget.setSizePolicy(
                    QSizePolicy(h_policy, v_policy,
                                QSizePolicy.ControlType.DefaultType)
                )

    def resizeEvent(self, event):
        self.resize_timer.start(200)
        super().resizeEvent(event)

    def handle_resize(self):
        """窗口缩放时自适应字体"""
        base_size = 18
        scale_h = self.height() / 1080
        new_size = max(12, int(base_size * scale_h))

        # 更新全局样式，但不影响特定组件
        # self.setStyleSheet(MACOS_STYLE.replace("18px", f"{new_size}px"))

        # 调整表格列宽
        if hasattr(self, 'steps_table'):
            total_width = self.steps_table.width()
            column_count = self.steps_table.columnCount()
            if column_count > 0:
                ratio = [0.1, 0.15, 0.6, 0.15]
                for i in range(column_count):
                    self.steps_table.setColumnWidth(i, int(total_width * ratio[i]))

    # ... (此处省略了大量未修改的 main.py 原有方法以缩减篇幅) ...
    # ... (init_manual_tab, init_auto_tab, send_manual_command, etc.) ...
    # ... (所有电机控制、自动化流程、预设管理的方法都保持不变) ...
    def init_manual_tab(self):
        layout = QVBoxLayout(self.manual_tab)
        layout.setContentsMargins(10, 10, 10, 10)

        # 微泵控制区
        motor_frame = QFrame()
        grid_layout = QGridLayout(motor_frame)
        self.motor_widgets = {}

        # 存储手动控制页的GroupBox引用，用于刷新标题
        self.manual_motor_groups = {}
        
        motors = [("X", 0, 0), ("Y", 0, 1), ("Z", 1, 0), ("A", 1, 1)]
        for motor, row, col in motors:
            # 获取备注并生成标题
            note = self.settings_manager.get_pump_note(motor)
            title = f"微泵 {motor} ({note})" if note else f"微泵 {motor}"
            group = QGroupBox(title)
            group.setFont(QFont("Microsoft YaHei", 16))
            self.manual_motor_groups[motor] = group
            
            # 启用右键菜单
            group.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
            group.customContextMenuRequested.connect(
                lambda pos, m=motor, g=group: self.show_pump_note_menu(m, pos, g)
            )
            
            vbox = QVBoxLayout(group)

            # 启用和持续开关
            switch_frame = QFrame()
            hbox = QHBoxLayout(switch_frame)
            enable_check = QCheckBox("启用")
            enable_check.setFont(QFont("Microsoft YaHei", 16))
            continuous_check = QCheckBox("持续")
            continuous_check.setFont(QFont("Microsoft YaHei", 16))
            hbox.addWidget(enable_check)
            hbox.addWidget(continuous_check)
            vbox.addWidget(switch_frame)

            # 方向选择
            dir_group = QButtonGroup(self)
            dir_frame = QFrame()
            hbox_dir = QHBoxLayout(dir_frame)
            forward_btn = QRadioButton("正转")
            backward_btn = QRadioButton("反转")
            forward_btn.setFont(QFont("Microsoft YaHei", 16))
            backward_btn.setFont(QFont("Microsoft YaHei", 16))
            dir_group.addButton(forward_btn)
            dir_group.addButton(backward_btn)
            forward_btn.setChecked(True)
            hbox_dir.addWidget(forward_btn)
            hbox_dir.addWidget(backward_btn)
            vbox.addWidget(dir_frame)

            # 参数输入
            speed_entry = QLineEdit()
            speed_entry.setPlaceholderText("速度值 (RPM)")
            speed_entry.setFont(QFont("Microsoft YaHei", 16))
            speed_entry.setFixedHeight(40)
            vbox.addWidget(speed_entry)

            angle_entry = QLineEdit()
            angle_entry.setPlaceholderText("角度")
            angle_entry.setFont(QFont("Microsoft YaHei", 16))
            angle_entry.setFixedHeight(40)
            vbox.addWidget(angle_entry)

            self.motor_widgets[motor] = {
                "enable": enable_check,
                "direction": dir_group,
                "speed": speed_entry,
                "angle": angle_entry,
                "continuous": continuous_check
            }

            grid_layout.addWidget(group, row, col)

        layout.addWidget(motor_frame)

        # 控制按钮区
        control_frame = QFrame()
        hbox = QHBoxLayout(control_frame)

        send_btn = QPushButton("发送指令")
        send_btn.setFont(QFont("Microsoft YaHei", 16))
        send_btn.setFixedHeight(40)
        send_btn.clicked.connect(self.send_manual_command)
        hbox.addWidget(send_btn)

        # 预设管理
        preset_frame = QFrame()
        preset_layout = QHBoxLayout(preset_frame)
        preset_lbl = QLabel("手动预设:")
        preset_lbl.setFont(QFont("Microsoft YaHei", 16))
        self.manual_preset_combo = QComboBox()
        self.manual_preset_combo.setFont(QFont("Microsoft YaHei", 16))
        self.manual_preset_combo.setFixedWidth(150)

        load_btn = QPushButton("加载")
        load_btn.setFont(QFont("Microsoft YaHei", 16))
        load_btn.clicked.connect(self.load_manual_preset)

        save_btn = QPushButton("保存")
        save_btn.setFont(QFont("Microsoft YaHei", 16))
        save_btn.clicked.connect(self.save_manual_preset)

        del_manual_btn = QPushButton("删除")
        del_manual_btn.clicked.connect(lambda: self.delete_preset('manual'))

        preset_layout.addWidget(preset_lbl)
        preset_layout.addWidget(self.manual_preset_combo)
        preset_layout.addWidget(load_btn)
        preset_layout.addWidget(save_btn)
        preset_layout.addWidget(del_manual_btn)
        hbox.addWidget(preset_frame)

        layout.addWidget(control_frame)

        # +++ 新增定时运行控件 +++
        timer_frame = QFrame()
        timer_layout = QHBoxLayout(timer_frame)

        # 定时运行标签
        timer_label = QLabel("定时运行:")
        timer_label.setFont(QFont("Microsoft YaHei", 16))

        # 时间输入框
        self.timer_input = QLineEdit()
        self.timer_input.setPlaceholderText("时长")
        self.timer_input.setFont(QFont("Microsoft YaHei", 16))
        self.timer_input.setFixedWidth(100)

        # 时间单位选择
        self.time_unit_combo = QComboBox()
        self.time_unit_combo.addItems(["秒", "分钟", "小时"])
        self.time_unit_combo.setFont(QFont("Microsoft YaHei", 16))
        self.time_unit_combo.setFixedWidth(100)

        # 按钮容器
        btn_container = QFrame()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)

        # 运行按钮
        self.timer_run_btn = QPushButton("运行")
        self.timer_run_btn.setFont(QFont("Microsoft YaHei", 16))
        self.timer_run_btn.clicked.connect(self.start_timed_run)

        # 新增按钮
        self.timer_pause_btn = QPushButton("暂停")
        self.timer_pause_btn.setFont(QFont("Microsoft YaHei", 16))
        self.timer_pause_btn.clicked.connect(self.pause_timed_run)
        self.timer_pause_btn.setEnabled(False)

        self.timer_resume_btn = QPushButton("继续")
        self.timer_resume_btn.setFont(QFont("Microsoft YaHei", 16))
        self.timer_resume_btn.clicked.connect(self.resume_timed_run)
        self.timer_resume_btn.setEnabled(False)

        self.timer_cancel_btn = QPushButton("取消")
        self.timer_cancel_btn.setFont(QFont("Microsoft YaHei", 16))
        self.timer_cancel_btn.clicked.connect(self.cancel_timed_run)
        self.timer_cancel_btn.setEnabled(False)
        btn_layout.addWidget(self.timer_run_btn)
        btn_layout.addWidget(self.timer_pause_btn)
        btn_layout.addWidget(self.timer_resume_btn)
        btn_layout.addWidget(self.timer_cancel_btn)

        timer_layout.addWidget(timer_label)
        timer_layout.addWidget(self.timer_input)
        timer_layout.addWidget(self.time_unit_combo)
        timer_layout.addWidget(btn_container)
        layout.addWidget(timer_frame)
        # 定时器相关初始化
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_timer)
        self.remaining_seconds = 0
        self.is_paused = False
        self.paused_seconds = 0

    def start_timed_run(self):
        # 验证输入
        try:
            duration = float(self.timer_input.text())
            unit = self.time_unit_combo.currentText()

            # 转换为秒
            if unit == "分钟":
                duration *= 60
            elif unit == "小时":
                duration *= 3600

            if duration <= 0:
                raise ValueError("时长必须大于0")

        except ValueError as e:
            QMessageBox.critical(self, "输入错误", str(e))
            return
        # 发送持续运行指令
        if not self.send_continuous_run_command():
            return
        # 启动定时器
        self.remaining_seconds = int(duration)
        self.timer.start(1000)  # 每秒更新
        self.timer_run_btn.setEnabled(False)
        self.update_status_message()
        # 更新按钮状态
        self.timer_run_btn.setEnabled(False)
        self.timer_pause_btn.setEnabled(True)
        self.timer_cancel_btn.setEnabled(True)
        self.timer_resume_btn.setEnabled(False)
        self.is_paused = False

    def send_continuous_run_command(self):
        # 生成持续运行指令（例如：XEFV100JG YEFV200JG）
        command = ""
        for motor in ["X", "Y", "Z", "A"]:
            widgets = self.motor_widgets[motor]
            if widgets["enable"].isChecked():
                speed = widgets["speed"].text()
                if not speed.isdigit():
                    QMessageBox.critical(self, "错误", f"电机{motor}速度值无效")
                    return False
                direction = "F" if widgets["direction"].checkedButton().text() == "正转" else "B"
                command += f"{motor}E{direction}V{speed}JG"

        if not command:
            QMessageBox.critical(self, "错误", "没有启用的电机")
            return False

        return self.send_command(command + "\r\n")

    def update_timer(self):
        if not self.is_paused:
            self.remaining_seconds -= 1
            if self.remaining_seconds <= 0:
                self.stop_timed_run()
                self.send_stop_command()
            self.update_status_message()

    def update_status_message(self):
        mins, secs = divmod(self.remaining_seconds, 60)
        hours, mins = divmod(mins, 60)
        time_str = f"{hours:02d}:{mins:02d}:{secs:02d}"
        self.status_bar.showMessage(f"运行中 - 剩余时间: {time_str}")

    def pause_timed_run(self):
        # 发送急停指令（示例指令，需根据实际协议调整）
        command = "".join([f"{motor}DFV0J0" for motor in ["X", "Y", "Z", "A"]])
        self.send_command(command + "\r\n")
        self.timer.stop()
        self.is_paused = True
        self.paused_seconds = self.remaining_seconds

        # 更新按钮状态
        self.timer_pause_btn.setEnabled(False)
        self.timer_resume_btn.setEnabled(True)
        self.status_bar.showMessage(f"运行已暂停，剩余时间: {self.format_time(self.paused_seconds)}")

    def resume_timed_run(self):
        if not self.send_continuous_run_command():
            return
        self.remaining_seconds = self.paused_seconds
        self.timer.start(1000)
        self.is_paused = False

        # 更新按钮状态
        self.timer_pause_btn.setEnabled(True)
        self.timer_resume_btn.setEnabled(False)
        self.update_status_message()

    def cancel_timed_run(self):
        self.timer.stop()
        self.send_stop_command()
        self.remaining_seconds = 0
        self.is_paused = False

        # 更新按钮状态
        self.timer_run_btn.setEnabled(True)
        self.timer_pause_btn.setEnabled(False)
        self.timer_resume_btn.setEnabled(False)
        self.timer_cancel_btn.setEnabled(False)
        self.status_bar.showMessage("运行已取消")

    def format_time(self, seconds):
        mins, secs = divmod(seconds, 60)
        hours, mins = divmod(mins, 60)
        return f"{hours:02d}:{mins:02d}:{secs:02d}"

    def stop_timed_run(self):
        self.timer.stop()
        self.timer_run_btn.setEnabled(True)
        self.timer_pause_btn.setEnabled(False)
        self.timer_resume_btn.setEnabled(False)
        self.timer_cancel_btn.setEnabled(False)
        self.status_bar.showMessage("定时运行完成")

    def send_stop_command(self):
        # 发送脱机指令（例如：XDFV0J0 YDFV0J0）
        command = "".join([f"{motor}DFV0J0" for motor in ["X", "Y", "Z", "A"]])
        self.send_command(command + "\r\n")

    def init_auto_tab(self):
        layout = QVBoxLayout(self.auto_tab)
        layout.setContentsMargins(8, 8, 8, 8)

        # 预设管理
        preset_frame = QFrame()
        hbox = QHBoxLayout(preset_frame)

        # 新增自动预设combo定义
        self.auto_preset_combo = QComboBox()
        self.auto_preset_combo.setFont(QFont("Microsoft YaHei", 16))
        self.auto_preset_combo.setFixedWidth(200)

        preset_lbl = QLabel("自动预设:")
        load_btn = QPushButton("加载")
        load_btn.clicked.connect(self.load_auto_preset)
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_auto_preset)
        del_auto_btn = QPushButton("删除")
        del_auto_btn.clicked.connect(lambda: self.delete_preset('auto'))

        # 将控件添加到布局
        hbox.addWidget(QLabel("自动预设:"))
        hbox.addWidget(self.auto_preset_combo)
        hbox.addWidget(load_btn)
        hbox.addWidget(save_btn)
        hbox.addWidget(del_auto_btn)
        layout.addWidget(preset_frame)

        # 步骤表格
        self.steps_table = DragDropTreeWidget()
        self.steps_table.setHeaderLabels(["编号", "名称", "参数配置", "间隔(s)"])
        self.steps_table.setRootIsDecorated(False)  # 隐藏根节点展开图标
        self.steps_table.setUniformRowHeights(True)  # 统一行高
        self.steps_table.setSortingEnabled(False)  # 禁用排序
        self.steps_table.setEditTriggers(QTreeWidget.EditTrigger.NoEditTriggers)  # 设置不可编辑
        for i in range(self.steps_table.columnCount()):
            header_item = self.steps_table.headerItem()
            header_item.setTextAlignment(i, Qt.AlignmentFlag.AlignCenter)
        self.steps_table.setColumnWidth(0, 80)
        self.steps_table.setColumnWidth(1, 150)
        self.steps_table.setColumnWidth(2, 500)
        self.steps_table.setColumnWidth(3, 120)
        self.steps_table.itemDoubleClicked.connect(self.edit_step)
        layout.addWidget(self.steps_table)

        # 控制按钮
        btn_frame = QFrame()
        hbox = QHBoxLayout(btn_frame)
        buttons = [
            ("添加步骤", self.add_step),
            ("删除步骤", self.remove_step),
            ("复制步骤", self.copy_step),
            ("粘贴步骤", self.paste_step),
            ("开始执行", self.start_automation),
            ("停止执行", self.stop_automation)
        ]
        for text, slot in buttons:
            btn = QPushButton(text)
            btn.setFont(QFont("Microsoft YaHei", 16))
            btn.setFixedHeight(40)
            btn.clicked.connect(slot)
            hbox.addWidget(btn)
        layout.addWidget(btn_frame)

        # 循环次数设置
        loop_frame = QFrame()
        hbox = QHBoxLayout(loop_frame)
        loop_label = QLabel("循环次数 (0=无限):")
        self.loop_entry = QLineEdit("1")
        self.loop_entry.setFixedWidth(100)
        hbox.addWidget(loop_label)
        hbox.addWidget(self.loop_entry)
        hbox.addStretch()
        layout.addWidget(loop_frame)

    def sync_automation_steps_order(self):
        """精确同步步骤顺序"""
        try:
            # 清空旧数据时保留有效项
            valid_items = [
                (self.steps_table.topLevelItem(i), i)
                for i in range(self.steps_table.topLevelItemCount())
                if self.steps_table.topLevelItem(i) is not None
            ]

            # 按显示顺序重建数据
            new_steps = []
            for idx, (item, _) in enumerate(valid_items, 1):
                if not item:
                    continue
                step_data = item.data(0, Qt.UserRole)
                if step_data:  # 有效性检查
                    new_steps.append(step_data)
                    item.setText(0, str(idx))  # 强制刷新编号

            # 重建数据内存
            self.automation_steps.clear()
            self.automation_steps.extend(new_steps)

            # 清除无效空白项并重新加载表格
            self.steps_table.clear()
            for step in self.automation_steps:
                self._add_step_to_table(step)

            self.log("步骤顺序已调整")
        except Exception as e:
            self.log(f"同步步骤顺序失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"同步步骤顺序失败: {str(e)}")

    def _add_step_to_table(self, step):
        """安全的表格项添加方法"""
        try:
            current_count = self.steps_table.topLevelItemCount()
            idx = current_count + 1

            # 参数解析逻辑
            params_desc = []
            for motor in ["X", "Y", "Z", "A"]:
                cfg = step.get(motor, {})
                if cfg.get("enable") == "E":
                    desc = f"{motor}:方向{cfg.get('direction', '?')} 速度{cfg.get('speed', '?')} 角度{cfg.get('angle', '?')}"
                    params_desc.append(desc)
            params_str = " | ".join(params_desc) if params_desc else "所有微泵脱机"
            interval_ms = step.get("interval", 0)
            name = step.get("name", f"步骤 {idx}")

            # 创建表格项
            item = QTreeWidgetItem([
                str(idx),
                name,
                params_str,
                f"{interval_ms / 1000.0:.1f}"
            ])
            item.setData(0, Qt.UserRole, step)

            # 确保属性设置
            for i in range(self.steps_table.columnCount()):
                item.setTextAlignment(i, Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsDropEnabled)  # 禁止作为拖放目标

            self.steps_table.addTopLevelItem(item)

        except Exception as e:
            self.log(f"步骤添加错误: {str(e)}")

    def edit_step(self, item, column):
        step_index = self.steps_table.indexOfTopLevelItem(item)
        if step_index >= 0 and step_index < len(self.automation_steps):
            step = self.automation_steps[step_index]
            config_window = MotorStepConfig(self, step_index + 1, step)
            if config_window.exec() == QDialog.DialogCode.Accepted:
                # 更新步骤参数
                self.automation_steps[step_index] = config_window.step_params
                self.update_steps_table()
                self.log(f"步骤 {step_index + 1} 已编辑")

    def delete_preset(self, preset_type):
        """删除预设的通用方法"""
        combo = self.manual_preset_combo if preset_type == 'manual' else self.auto_preset_combo
        preset_name = combo.currentText()
        if not preset_name:
            QMessageBox.warning(self, "警告", "请先选择要删除的预设")
            return
        # 确认对话框
        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除预设 '{preset_name}' 吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            full_name = f"{preset_type}_{preset_name}"
            if full_name in self.presets:
                del self.presets[full_name]
                self._preset_manager.save_all()
                self.update_preset_combos()
                self.log(f"预设 '{preset_name}' 已删除")
            else:
                QMessageBox.critical(self, "错误", "找不到指定的预设")

    # 新增步骤复制粘贴功能
    def copy_step(self):
        selected = self.steps_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "警告", "请先选择要复制的步骤")
            return

        item = selected[0]
        step_index = self.steps_table.indexOfTopLevelItem(item)
        if 0 <= step_index < len(self.automation_steps):
            self.copied_step = self.automation_steps[step_index].copy()
            self.log(f"已复制步骤 {step_index + 1}")

    def paste_step(self):
        if not self.copied_step:
            QMessageBox.warning(self, "警告", "请先复制一个步骤")
            return
        # 获取插入位置
        selected = self.steps_table.selectedItems()
        insert_index = len(self.automation_steps)  # 默认插入末尾
        if selected:
            insert_index = self.steps_table.indexOfTopLevelItem(selected[0])
        # 创建新步骤并插入
        new_step = self.copied_step.copy()
        new_step['name'] = f"{new_step.get('name', '步骤')} (副本)"
        self.automation_steps.insert(insert_index, new_step)
        self.update_steps_table()
        self.log(f"已粘贴步骤到位置 {insert_index + 1}")

    def generate_command(self, step_params):
        """生成带校准的指令（优化版：支持 PID 闭环模式）"""
        command = ""
        command_active_motors = set()
        self.pending_targets = {m: None for m in ["X", "Y", "Z", "A"]}
        self.expected_rotation = {m: 0 for m in ["X", "Y", "Z", "A"]}
        direction_map = {"F": 1, "B": -1}

        # 自动模式初始基准设置 (逻辑保持不变)
        if self.running_mode == "auto" and self.is_first_command:
            for motor in self.active_motors:
                if self.current_angles[motor] is not None:
                    self.initial_angle_base[motor] = self.current_angles[motor]
                else:
                    self.initial_angle_base[motor] = None

        for motor in ["X", "Y", "Z", "A"]:
            config = step_params.get(motor, {})
            enable = config.get("enable", "D")

            # 核心修改：如果电机未启用，则直接跳过，不生成任何指令
            if enable != "E":
                continue

            # 如果程序执行到这里，说明电机已启用
            command_active_motors.add(motor)

            direction = config.get("direction", "F")
            speed = config.get("speed", "0")
            raw_angle = config.get("angle", "0").upper()
            is_continuous = config.get("continuous", False)
            dir_factor = direction_map[direction]

            try:
                # 连续转动模式：始终使用传统开环指令
                if is_continuous:
                    command += f"{motor}E{direction}V{speed}JG"
                    self.pending_targets[motor] = None
                    continue

                raw_rotation = float(raw_angle)
                self.expected_rotation[motor] = raw_rotation

                # ===== PID 精确控制模式 =====
                # 条件：自动校准开启 且 非连续模式
                if self.auto_calibration_enabled and not is_continuous:
                    current = self.current_angles.get(motor, 0.0)
                    precision = getattr(self, 'pid_precision', 0.5)
                    
                    if self.running_mode == "auto":
                        # 自动模式：使用理论目标角度（基于初始基准 + 累积旋转）
                        # 这样可以补偿累积误差，确保长期运行的精度
                        if self.initial_angle_base[motor] is None:
                            self.initial_angle_base[motor] = current
                        
                        # 先更新累积旋转量
                        self.accumulated_rotation[motor] += raw_rotation * dir_factor
                        
                        # 计算理论目标角度（基于初始基准）
                        target_angle = (self.initial_angle_base[motor] + self.accumulated_rotation[motor]) % 360
                        self.expected_angles[motor] = target_angle
                    else:
                        # 手动模式：基于当前角度计算目标
                        target_angle = (current + raw_rotation * dir_factor) % 360
                    
                    # 生成 PID 定位指令: XEFT90.0P0.5
                    command += f"{motor}E{direction}T{target_angle:.1f}P{precision}"
                    
                    # 更新期望目标
                    self.pending_targets[motor] = target_angle
                    
                    continue

                # ===== 传统开环模式 =====
                if self.running_mode == "auto":
                    if self.initial_angle_base[motor] is None:
                        base = self.current_angles.get(motor, 0.0)
                        self.initial_angle_base[motor] = base

                    raw_rotation_signed = float(raw_angle) * dir_factor
                    self.accumulated_rotation[motor] += raw_rotation_signed

                    # 传统自动校准补偿（当 PID 模式关闭时）
                    calibrated_rotation = raw_rotation_signed

                    actual_rotation = abs(calibrated_rotation)
                    self.expected_angles[motor] = (
                            (self.initial_angle_base[motor] + self.accumulated_rotation[motor]) % 360)

                else:  # 手动模式
                    actual_rotation = float(raw_angle)
                    current = self.current_angles.get(motor, 0.0)
                    self.pending_targets[motor] = (current + actual_rotation * dir_factor) % 360

                # 将当前启用电机的指令拼接到总指令字符串中
                command += f"{motor}E{direction}V{speed}J{actual_rotation:.3f}"

            except (ValueError, TypeError):
                self.log(f"电机{motor}参数错误，已跳过: 速度='{speed}', 角度='{raw_angle}'")
                # 如果参数解析错误，则记录日志并跳过此电机，不影响其他电机
                continue

        # 更新活动电机状态，并附加结束符
        self.active_motors = command_active_motors
        self.is_first_command = False

        # 只有在至少一个电机被启用时才添加回车换行符
        if command:
            return command + "\r\n"
        return ""  # 如果没有启用的电机，则返回空字符串，不发送任何指令

    def get_available_ports(self):
        # 固定添加COM1和COM2
        ports = {port.device for port in list_ports.comports()}
        fixed_ports = {"COM1", "COM2"}
        return sorted(fixed_ports.union(ports), key=lambda x: int(x[3:]) if x.startswith("COM") else 999)

    def refresh_serial_ports(self):
        current = self.port_combo.currentText()
        current_baud = self.baud_combo.currentText()
        self.port_combo.clear()
        available_ports = self.get_available_ports()
        self.port_combo.addItems(available_ports)
        self.baud_combo.setCurrentText(current_baud)
        # 检查当前端口是否在新的可用端口中
        if current in available_ports:
            self.port_combo.setCurrentText(current)
        else:
            if available_ports:
                self.port_combo.setCurrentIndex(0)

    def toggle_serial(self):
        if self.serial_port and self.serial_port.is_open:
            self.close_serial()
        else:
            self.open_serial()

    def open_serial(self):
        try:
            port = self.port_combo.currentText()
            baudrate = int(self.baud_combo.currentText())
            if not port:
                raise serial.SerialException("未选择端口")

            self.serial_port = serial.Serial(
                port=port,
                baudrate=baudrate,
                timeout=1,
                write_timeout=1
            )
            self.serial_port.reset_input_buffer()
            self.serial_port.reset_output_buffer()

            self._closing = False  # 重置关闭标志
            self.serial_reader = SerialReader(self.serial_port)
            # 使用 QueuedConnection 确保信号在主线程处理
            self.serial_reader.data_received.connect(
                self.handle_serial_data, Qt.ConnectionType.QueuedConnection)
            self.serial_reader.pid_packet_received.connect(
                self.handle_pid_packet, Qt.ConnectionType.QueuedConnection)
            self.serial_reader.test_result_received.connect(
                self.handle_test_result_packet, Qt.ConnectionType.QueuedConnection)
            self.serial_reader.angle_packet_received.connect(
                self.handle_angle_packet, Qt.ConnectionType.QueuedConnection)
            self.serial_reader.start()
            
            # 启动图表批量更新定时器
            self._chart_update_timer.start()

            self.connect_btn.setText("关闭串口")
            self.log(f"串口已连接 {port}@{baudrate}")
            self.status_bar.showMessage(f"已连接 {port}@{baudrate}")
        except serial.SerialException as e:
            QMessageBox.critical(self, "串口错误", f"串口连接失败: {e}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"发生未知错误: {e}")

    def close_serial(self):
        # 设置关闭标志，阻止新的信号处理
        self._closing = True
        
        # 停止图表批量更新定时器
        try:
            if hasattr(self, '_chart_update_timer') and self._chart_update_timer.isActive():
                self._chart_update_timer.stop()
        except (RuntimeError, AttributeError):
            pass
        
        # 清空待处理的数据包
        if hasattr(self, '_pending_pid_packets'):
            self._pending_pid_packets.clear()
        
        # 先停止 PID 更新定时器
        try:
            if hasattr(self, 'pid_update_timer') and self.pid_update_timer.isActive():
                self.pid_update_timer.stop()
        except (RuntimeError, AttributeError):
            pass
        
        # 发送停止角度流指令到下位机
        with self.serial_lock:
            if self.serial_port and self.serial_port.is_open:
                try:
                    self.serial_port.write(b"ANGLESTREAM_STOP\r\n")
                    self.serial_port.flush()
                    time.sleep(0.1)  # 等待下位机处理
                except Exception as e:
                    if not self._closing:
                        self.log(f"发送停止角度流指令失败: {e}")
        
        # 在锁外先断开信号连接（避免死锁）
        reader = None
        if hasattr(self, 'serial_reader') and self.serial_reader is not None:
            reader = self.serial_reader
            self.serial_reader = None  # 先置空，防止信号处理
            
            # 断开所有信号连接
            try:
                reader.data_received.disconnect()
            except (TypeError, RuntimeError):
                pass
            try:
                reader.pid_packet_received.disconnect()
            except (TypeError, RuntimeError):
                pass
            try:
                reader.test_result_received.disconnect()
            except (TypeError, RuntimeError):
                pass
            try:
                reader.angle_packet_received.disconnect()
            except (TypeError, RuntimeError):
                pass
        
        # 停止线程（在锁外）
        if reader is not None:
            reader.stop()  # stop方法内部会等待
        
        # 关闭串口
        with self.serial_lock:
            if self.serial_port and self.serial_port.is_open:
                try:
                    self.serial_port.close()
                except Exception:
                    pass
                self.serial_port = None
        
        self.connect_btn.setText("打开串口")
        self.log("串口已关闭")
        self.status_bar.showMessage("串口已关闭")

    def send_manual_command(self):
        params = {}
        active_motors = set()
        self.running_mode = "manual"
        for motor, widgets in self.motor_widgets.items():
            params[motor] = {
                "enable": "E" if widgets["enable"].isChecked() else "D",
                "direction": "F" if widgets["direction"].checkedButton().text() == "正转" else "B",
                "speed": widgets["speed"].text(),
                "angle": widgets["angle"].text().upper(),
                "continuous": widgets["continuous"].isChecked()
            }
            if params[motor]["enable"] == "E":
                active_motors.add(motor)
            # 验证输入
            try:
                if params[motor]["speed"]: float(params[motor]["speed"])
                if params[motor]["angle"] and params[motor]["angle"] != "G": float(params[motor]["angle"])
            except ValueError:
                QMessageBox.critical(self, "输入错误", f"电机 {motor} 的速度或角度值无效")
                return

        self.active_motors = active_motors
        command = self.generate_command(params)
        if self.send_command(command):
            self.status_bar.showMessage("手动指令已发送")

    def send_command(self, command):
        if not self.serial_port or not self.serial_port.is_open:
            QMessageBox.critical(self, "错误", "请先打开串口连接！")
            return False
        try:
            with self.serial_lock:
                self.serial_port.write(command.encode('utf-8'))
            self.log(f"已发送指令: {command.strip()}")
            return True
        except Exception as e:
            QMessageBox.critical(self, "发送错误", str(e))
            self.close_serial()
            return False

    def add_step(self):
        try:
            step_num = len(self.automation_steps) + 1
            config_window = MotorStepConfig(self, step_num)
            if config_window.exec() == QDialog.DialogCode.Accepted:
                if config_window.step_params:
                    self.automation_steps.append(config_window.step_params)
                    self.update_steps_table()
                    self.log(f"已成功添加步骤 {step_num}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"添加步骤失败: {str(e)}")

    def remove_step(self):
        selected = self.steps_table.selectedItems()
        if selected:
            index = self.steps_table.indexOfTopLevelItem(selected[0])
            if 0 <= index < len(self.automation_steps):
                del self.automation_steps[index]
                self.update_steps_table()
                self.log(f"已删除步骤 {index + 1}")

    def start_automation(self):
        # 前置检查
        if not hasattr(self, '_automation_mutex'):
            self._automation_mutex = threading.Lock()

        with self._automation_mutex:
            if self.automation_thread and self.automation_thread.isRunning():
                QMessageBox.warning(self, "警告", "已有正在运行的自动化任务")
                return

            if not self.serial_port or not self.serial_port.is_open:
                QMessageBox.critical(self, "错误", "串口未连接")
                return

            try:
                self.loop_count = max(0, int(self.loop_entry.text()))
            except ValueError:
                QMessageBox.critical(self, "错误", "无效的循环次数")
                return

            self.running_mode = "auto"
            self.is_first_command = True
            for motor in ["X", "Y", "Z", "A"]:
                self.initial_angle_base[motor] = self.current_angles.get(motor)
                self.accumulated_rotation[motor] = 0.0

            self.automation_thread = AutomationThread(
                weakref.ref(self),
                self.automation_steps,
                self.loop_count,
                self.serial_port,
                self.serial_lock
            )
            self.automation_thread.update_status.connect(self.status_bar.showMessage)
            self.automation_thread.error_occurred.connect(self.handle_automation_error)
            self.automation_thread.finished.connect(self._on_automation_finished)
            self.automation_thread.start()
            self.log("自动化任务安全启动")

    def _on_automation_finished(self):
        # 使用QTimer延迟清理，确保信号处理完成
        from PySide6.QtCore import QTimer
        QTimer.singleShot(100, self._cleanup_automation_thread)
        self.status_bar.showMessage("自动化运行已完成")
        self.log("自动化任务已完成")
    
    def _cleanup_automation_thread(self):
        """延迟清理自动化线程"""
        # 确保互斥锁存在
        if not hasattr(self, '_automation_mutex'):
            self._automation_mutex = threading.Lock()
        
        with self._automation_mutex:
            if self.automation_thread is not None:
                thread = self.automation_thread
                try:
                    thread.update_status.disconnect()
                    thread.error_occurred.disconnect()
                    thread.finished.disconnect()
                except (TypeError, RuntimeError):
                    pass
                self.automation_thread = None

    def handle_automation_error(self, message):
        # 先记录日志，再停止自动化，避免在停止过程中出现问题
        self.log(f"错误: {message}")
        # 使用QTimer延迟停止，避免在信号处理中直接停止线程
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, lambda: self._handle_automation_error_delayed(message))
    
    def _handle_automation_error_delayed(self, message):
        """延迟处理自动化错误"""
        self.stop_automation()
        QMessageBox.critical(self, "运行错误", message)

    def stop_automation(self):
        # 确保互斥锁存在
        if not hasattr(self, '_automation_mutex'):
            self._automation_mutex = threading.Lock()
        
        # 先标记停止状态，防止其他操作
        self.running = False
        
        # 清空待处理的数据包
        if hasattr(self, '_pending_pid_packets'):
            self._pending_pid_packets.clear()
        
        # 停止 PID 更新定时器
        try:
            if hasattr(self, 'pid_update_timer') and self.pid_update_timer.isActive():
                self.pid_update_timer.stop()
        except (RuntimeError, AttributeError):
            pass
        
        # 停止 PID 分析器中的所有活动记录
        try:
            if hasattr(self, 'pid_analyzer') and self.pid_analyzer is not None:
                self.pid_analyzer.stop_all()
        except (RuntimeError, AttributeError):
            pass
        
        # 重置状态面板
        try:
            if hasattr(self, 'pid_stats_panel') and self.pid_stats_panel is not None:
                self.pid_stats_panel.reset_all()
        except (RuntimeError, AttributeError):
            pass
        
        # 先断开自动化线程的信号连接（在锁外）
        thread = None
        if self.automation_thread is not None:
            thread = self.automation_thread
            self.automation_thread = None  # 先置空，防止重入
            
            # 断开信号连接
            try:
                thread.update_status.disconnect()
            except (TypeError, RuntimeError):
                pass
            try:
                thread.error_occurred.disconnect()
            except (TypeError, RuntimeError):
                pass
            try:
                thread.finished.disconnect()
            except (TypeError, RuntimeError):
                pass
        
        # 停止线程（在锁外，避免死锁）
        if thread is not None and thread.isRunning():
            thread.safe_stop()
            # 等待线程完全结束，最多等待2秒
            if not thread.wait(2000):
                self.log("警告: 自动化线程未能在2秒内停止")
        
        # 发送停止指令给下位机（停止所有电机和 PID 模式）
        try:
            if self.serial_port and self.serial_port.is_open:
                # 先停止 PID 定位模式
                self.send_command("PIDSTOP\r\n")
                # 再停止传统开环模式
                stop_cmd = "XDFV0J0YDFV0J0ZDFV0J0ADFV0J0\r\n"
                self.send_command(stop_cmd)
        except Exception as e:
            self.log(f"发送停止指令失败: {e}")
        
        self.log("自动化运行已停止")
        self.status_bar.showMessage("自动化运行已停止")

    def update_steps_table(self):
        scroll_pos = self.steps_table.verticalScrollBar().value()
        self.steps_table.clear()
        for step in self.automation_steps:
            self._add_step_to_table(step)
        self.steps_table.verticalScrollBar().setValue(scroll_pos)

    def _update_step_item(self, item, step, idx):
        params_desc = []
        for motor in ["X", "Y", "Z", "A"]:
            cfg = step.get(motor, {})
            if cfg.get("enable") == "E":
                desc = f"{motor}:方向{cfg.get('direction', '?')} 速度{cfg.get('speed', '?')} 角度{cfg.get('angle', '?')}"
                params_desc.append(desc)
        params_str = " | ".join(params_desc) if params_desc else "所有微泵脱机"
        item.setText(0, str(idx))
        item.setText(1, step.get("name", f"步骤 {idx}"))
        item.setText(2, params_str)
        # 显示为秒
        interval_ms = step.get("interval", 0)
        item.setText(3, f"{interval_ms / 1000.0:.1f}")
        item.setData(0, Qt.UserRole, step)

    def save_manual_preset(self):
        preset_name, ok = QInputDialog.getText(self, "保存手动预设", "输入预设名称:")
        if not (ok and preset_name.strip()): return
        manual_params = {}
        for motor, widgets in self.motor_widgets.items():
            manual_params[motor] = {
                "enable": "E" if widgets["enable"].isChecked() else "D",
                "direction": "F" if widgets["direction"].checkedButton().text() == "正转" else "B",
                "speed": widgets["speed"].text(),
                "angle": widgets["angle"].text(),
                "continuous": widgets["continuous"].isChecked()
            }
        self.presets[f"manual_{preset_name}"] = manual_params
        self._preset_manager.save_all()
        self.update_preset_combos()
        self.log(f"手动预设 '{preset_name}' 已保存")

    def load_manual_preset(self):
        name = self.manual_preset_combo.currentText()
        if not name: return
        preset_data = self.presets.get(f"manual_{name}")
        if preset_data:
            for motor, params in preset_data.items():
                if motor in self.motor_widgets:
                    widgets = self.motor_widgets[motor]
                    widgets["enable"].setChecked(params.get("enable") == "E")
                    if params.get("direction") == "F":
                        widgets["direction"].buttons()[0].setChecked(True)
                    else:
                        widgets["direction"].buttons()[1].setChecked(True)
                    widgets["speed"].setText(params.get("speed", ""))
                    widgets["angle"].setText(params.get("angle", ""))
                    widgets["continuous"].setChecked(params.get("continuous", False))
            self.log(f"已加载手动预设 '{name}'")

    def save_auto_preset(self):
        preset_name, ok = QInputDialog.getText(self, "保存自动预设", "输入预设名称:")
        if not (ok and preset_name.strip()): return
        preset_data = {
            "steps": [s.copy() for s in self.automation_steps],
            "loop_count": self.loop_entry.text()
        }
        self.presets[f"auto_{preset_name}"] = preset_data
        self._preset_manager.save_all()
        self.update_preset_combos()
        self.log(f"自动预设 '{preset_name}' 已保存")

    def load_auto_preset(self):
        preset_name = self.auto_preset_combo.currentText()
        if not preset_name: return
        preset_data = self.presets.get(f"auto_{preset_name}")
        if preset_data:
            self.automation_steps = [step.copy() for step in preset_data.get("steps", [])]
            self.loop_entry.setText(str(preset_data.get("loop_count", 1)))
            self.update_steps_table()
            self.log(f"已加载自动预设 '{preset_name}'")

    # =================================================================
    # --- 新增功能：自动保存与加载设置 ---
    # =================================================================
    def save_settings(self):
        """将当前UI设置保存到JSON文件。"""
        # 使用settings_manager的内部设置，确保保留motor配置
        # 先重新加载文件以获取最新的motor配置
        self.settings_manager.load()
        
        # 更新UI相关设置到settings_manager
        self.settings_manager.set_section("serial", {
            "port": self.port_combo.currentText(),
            "baudrate": self.baud_combo.currentText()
        })
        self.settings_manager.set_section("window", {
            "width": self.size().width(),
            "height": self.size().height()
        })
        self.settings_manager.set_section("calibration", {
            "enabled": self.auto_cal_switch.isChecked(),
            "amplitude": self.calibration_amp_input.text()
        })
        
        # 仅当光谱仪库可用时才保存相关设置
        if NIDAQMX_AVAILABLE:
            self.settings_manager.set_section("spectrometer", {
                "device": self.spectro_device_combo.currentText(),
                "channel": self.spectro_channel_combo.currentText(),
                "rate": self.spectro_rate_spin.value()
            })

        # 使用settings_manager保存，确保所有配置（包括motor）都被保存
        if self.settings_manager.save():
            self.log("设置已成功保存。")
        else:
            self.log("保存设置时出错。")

    def load_settings(self):
        """从JSON文件加载UI设置。"""
        if not os.path.exists(self.settings_file):
            self.log("未找到设置文件，使用默认值。")
            return

        try:
            with open(self.settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)

            # 加载串口设置
            serial_settings = settings.get("serial", {})
            if "port" in serial_settings:
                saved_port = serial_settings["port"]
                available_ports = [self.port_combo.itemText(i) for i in range(self.port_combo.count())]
                if saved_port in available_ports:
                    self.port_combo.setCurrentText(saved_port)
                elif available_ports:
                    self.port_combo.setCurrentIndex(0)  # 智能选择第一个
            if "baudrate" in serial_settings:
                self.baud_combo.setCurrentText(serial_settings["baudrate"])

            # 加载窗口尺寸
            window_settings = settings.get("window", {})
            if "width" in window_settings and "height" in window_settings:
                self.resize(window_settings["width"], window_settings["height"])

            # 加载自动校准设置
            cal_settings = settings.get("calibration", {})
            if "enabled" in cal_settings:
                self.auto_cal_switch.setChecked(cal_settings["enabled"])
            if "amplitude" in cal_settings:
                self.calibration_amp_input.setText(cal_settings["amplitude"])

            # 加载光谱仪设置
            if NIDAQMX_AVAILABLE:
                spectro_settings = settings.get("spectrometer", {})
                if "device" in spectro_settings:
                    saved_device = spectro_settings["device"]
                    available_devices = [self.spectro_device_combo.itemText(i) for i in
                                         range(self.spectro_device_combo.count())]
                    if saved_device in available_devices:
                        self.spectro_device_combo.setCurrentText(saved_device)

                QApplication.processEvents()  # 等待UI响应设备更改，以便通道列表刷新

                if "channel" in spectro_settings:
                    saved_channel = spectro_settings["channel"]
                    available_channels = [self.spectro_channel_combo.itemText(i) for i in
                                          range(self.spectro_channel_combo.count())]
                    if saved_channel in available_channels:
                        self.spectro_channel_combo.setCurrentText(saved_channel)

                if "rate" in spectro_settings:
                    self.spectro_rate_spin.setValue(spectro_settings["rate"])

            # 加载零点偏移量（使用已初始化的settings_manager）
            self.angle_offsets = self.settings_manager.get_angle_offsets()
            self._update_offset_labels()  # 更新UI显示
            self.log(f"零点偏移量已加载: {self.angle_offsets}")

            self.log("设置已成功加载。")
        except Exception as e:
            self.log(f"加载设置时出错: {str(e)}。将使用默认值。")

    def update_preset_combos(self):
        manual_presets = [k[7:] for k in self.presets if k.startswith("manual_")]
        self.manual_preset_combo.clear()
        self.manual_preset_combo.addItems(sorted(manual_presets))

        auto_presets = [k[5:] for k in self.presets if k.startswith("auto_")]
        self.auto_preset_combo.clear()
        self.auto_preset_combo.addItems(sorted(auto_presets))

    def log(self, message):
        self.log_signal.emit(message)

    def _log_impl(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        self.log_text.append(f"[{timestamp}] {message}")
        self.log_text.moveCursor(QTextCursor.End)

    def clear_log(self):
        self.log_text.clear()

    def closeEvent(self, event):
        # 设置关闭标志，阻止所有信号处理
        self._closing = True
        
        # --- 保存设置 ---
        try:
            self.save_settings()
        except Exception:
            pass
        
        # 停止所有定时器
        try:
            if hasattr(self, '_chart_update_timer'):
                self._chart_update_timer.stop()
            if hasattr(self, 'pid_update_timer'):
                self.pid_update_timer.stop()
            if hasattr(self, 'resize_timer'):
                self.resize_timer.stop()
            if hasattr(self, 'spectro_timer'):
                self.spectro_timer.stop()
        except (RuntimeError, AttributeError):
            pass
        
        # 清空数据缓冲区
        if hasattr(self, '_pending_pid_packets'):
            self._pending_pid_packets.clear()

        # 统一的关闭清理
        try:
            self.stop_automation()
        except Exception:
            pass
        
        if NIDAQMX_AVAILABLE:
            try:
                self._spectro_stop_measurement()
            except Exception:
                pass
        
        try:
            self.close_serial()
        except Exception:
            pass

        if platform.system() == 'Windows':
            try:
                windll.winmm.timeEndPeriod(1)
            except Exception:
                pass

        event.accept()


    # ============= 零点标定功能 =============
    
    def set_current_as_zero(self, motor: str) -> None:
        """将当前角度设为零点"""
        if motor not in self.raw_angles:
            self.log(f"微泵{motor}原始角度数据不可用")
            return
        
        # 使用原始物理角度作为偏移量
        offset = self.raw_angles[motor]
        self.angle_offsets[motor] = offset
        
        # 保存到配置
        self.settings_manager.set_angle_offset(motor, offset)
        
        # 立即刷新显示（当前角度变为0）
        self.current_angles[motor] = 0.0
        self._update_angle_displays()
        self._update_offset_labels()
        
        # 日志输出
        self.log(f"微泵{motor}零点已设置: 物理角度={offset:.2f}°")
    
    def reset_zero_offsets(self) -> None:
        """重置所有零点偏移"""
        self.angle_offsets = {"X": 0.0, "Y": 0.0, "Z": 0.0, "A": 0.0}
        
        # 保存到配置
        self.settings_manager.reset_angle_offsets()
        
        # 恢复为原始角度
        for motor in ["X", "Y", "Z", "A"]:
            self.current_angles[motor] = self.raw_angles[motor]
        
        self._update_angle_displays()
        self._update_offset_labels()
        self.log("所有微泵零点已重置")
    
    def reset_deviation_data(self) -> None:
        """重置偏差计算相关数据"""
        # 重置理论偏差相关
        self.initial_angle_base = {m: None for m in ["X", "Y", "Z", "A"]}
        self.accumulated_rotation = {m: 0.0 for m in ["X", "Y", "Z", "A"]}
        self.theoretical_deviations = {m: None for m in ["X", "Y", "Z", "A"]}
        self.theoretical_target = {m: None for m in ["X", "Y", "Z", "A"]}
        
        # 重置实时偏差相关
        self.pending_targets = {}
        self.expected_rotation = {"X": 0, "Y": 0, "Z": 0, "A": 0}
        self.expected_changes = {}
        
        # 构造数据并调用现有方法刷新表格
        data = {
            "current": {m: self.current_angles.get(m, 0.0) for m in ["X", "Y", "Z", "A"]},
            "targets": {m: None for m in ["X", "Y", "Z", "A"]},
            "theoretical": {m: None for m in ["X", "Y", "Z", "A"]},
            "realtime": {m: None for m in ["X", "Y", "Z", "A"]}
        }
        self.update_angles(data)
        
        self.log("偏差数据已重置")
    
    def _update_angle_displays(self) -> None:
        """更新所有角度相关显示"""
        # 触发现有的角度更新机制
        current_angles = {m: self.current_angles.get(m, 0.0) for m in ["X", "Y", "Z", "A"]}
        
        # 更新电机圆圈动画和角度标签
        for motor in ["X", "Y", "Z", "A"]:
            if motor in self.motors:
                angle = current_angles[motor]
                self.motors[motor].set_angle(-angle)  # 负号用于顺时针旋转
            if motor in self.angle_labels:
                self.angle_labels[motor].setText(f"{current_angles[motor]:.3f}°")
    
    def _update_offset_labels(self) -> None:
        """更新偏移量显示标签"""
        if not hasattr(self, 'offset_labels'):
            return
        
        for motor in ["X", "Y", "Z", "A"]:
            offset = self.angle_offsets.get(motor, 0.0)
            if motor in self.offset_labels:
                self.offset_labels[motor].setText(f"偏移: {offset:.2f}°")
                # 偏移量为0时显示灰色，非零时显示蓝色
                color = "#8e8e93" if offset == 0.0 else "#007aff"
                self.offset_labels[motor].setStyleSheet(f"color: {color}; font-size: 13px;")

    # ============= 微泵备注功能 =============
    
    def show_pump_note_menu(self, motor: str, pos, group=None) -> None:
        """显示微泵备注右键菜单"""
        menu = QMenu(self)
        
        edit_action = menu.addAction("编辑备注")
        edit_action.triggered.connect(lambda: self.edit_pump_note(motor))
        
        current_note = self.settings_manager.get_pump_note(motor)
        if current_note:
            clear_action = menu.addAction("清除备注")
            clear_action.triggered.connect(lambda: self.clear_pump_note(motor))
        
        # 使用传入的group或sender来映射坐标
        widget = group if group else self.sender()
        if widget:
            menu.exec(widget.mapToGlobal(pos))
    
    def edit_pump_note(self, motor: str) -> None:
        """编辑微泵备注"""
        current_note = self.settings_manager.get_pump_note(motor)
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"微泵 {motor} 备注")
        dialog.setFixedSize(300, 140)
        
        layout = QVBoxLayout(dialog)
        
        # 输入框
        note_input = QLineEdit(current_note)
        note_input.setMaxLength(8)
        note_input.setPlaceholderText("输入备注(最多8字)")
        note_input.setFont(QFont("Microsoft YaHei", 14))
        layout.addWidget(note_input)
        
        # 提示
        hint = QLabel("限制8个字符")
        hint.setStyleSheet("color: #888; font-size: 12px;")
        layout.addWidget(hint)
        
        # 按钮
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(dialog.accept)
        clear_btn = QPushButton("清除")
        clear_btn.clicked.connect(lambda: note_input.clear())
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(clear_btn)
        layout.addLayout(btn_layout)
        
        if dialog.exec() == QDialog.Accepted:
            new_note = note_input.text().strip()
            self.settings_manager.set_pump_note(motor, new_note)
            self.refresh_pump_titles()
            self.log(f"微泵{motor}备注已更新: {new_note if new_note else '(已清除)'}")
    
    def clear_pump_note(self, motor: str) -> None:
        """清除微泵备注"""
        self.settings_manager.clear_pump_note(motor)
        self.refresh_pump_titles()
        self.log(f"微泵{motor}备注已清除")
    
    def refresh_pump_titles(self) -> None:
        """刷新所有微泵卡片标题"""
        for motor in ["X", "Y", "Z", "A"]:
            note = self.settings_manager.get_pump_note(motor)
            title = f"微泵 {motor} ({note})" if note else f"微泵 {motor}"
            
            # 更新转子监控页的卡片标题
            if hasattr(self, 'position_motor_groups') and motor in self.position_motor_groups:
                self.position_motor_groups[motor].setTitle(title)
            
            # 更新手动控制页的卡片标题
            if hasattr(self, 'manual_motor_groups') and motor in self.manual_motor_groups:
                self.manual_motor_groups[motor].setTitle(title)
    
    def get_pump_title(self, motor: str) -> str:
        """获取带备注的微泵标题"""
        note = self.settings_manager.get_pump_note(motor)
        return f"微泵 {motor} ({note})" if note else f"微泵 {motor}"


