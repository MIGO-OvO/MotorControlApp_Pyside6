"""基线稳定测试独立页面 Mixin。"""

from __future__ import annotations

import os
import time
from datetime import datetime
from typing import Optional

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QVBoxLayout,
)

from src.config.constants import (
    BUTTON_DANGER,
    BUTTON_SECONDARY,
    BUTTON_SUCCESS,
    COLOR_ACCENT_BLUE,
    COLOR_ACCENT_GREEN,
    COLOR_ACCENT_ORANGE,
    COLOR_BORDER_LIGHT,
    COLOR_BG_SUBTLE,
    COLOR_PRIMARY,
    COLOR_TEXT_MUTED,
    COLOR_TEXT_PRIMARY,
    COLOR_TEXT_SECONDARY,
    DEFAULT_BASELINE_DURATION_MIN,
    DEFAULT_BASELINE_WARMUP_S,
    DEFAULT_SPECTRO_CHART_POINTS,
    SPECTRO_BASELINE_DURATION_RANGE,
    SPECTRO_BASELINE_WARMUP_RANGE,
    SPECTRO_CHART_POINTS_RANGE,
    SPECTRO_CHART_POINTS_STEP,
)
from src.core.spectro_baseline import (
    BaselineExportRecord,
    BaselineMetrics,
    BaselineSample,
    analyze_baseline,
    build_baseline_export_tables,
)

try:
    import pyqtgraph as pg

    PYQTGRAPH_AVAILABLE = True
except ImportError:
    PYQTGRAPH_AVAILABLE = False
    pg = None  # type: ignore


class BaselineMixin:
    """基线稳定测试页面和逻辑。"""

    def _baseline_init_vars(self) -> None:
        self.baseline_records: list[BaselineExportRecord] = []
        self.baseline_voltage_data: list[float] = []
        self.baseline_metrics: Optional[BaselineMetrics] = None
        self.baseline_is_running: bool = False
        self.baseline_owned_measurement: bool = False
        self.baseline_start_time: float = 0.0
        self.baseline_end_time: float = 0.0
        self.baseline_started_at: Optional[datetime] = None
        self.baseline_finished_at: Optional[datetime] = None
        self.baseline_max_data_points: int = DEFAULT_SPECTRO_CHART_POINTS
        self.baseline_finish_timer = QTimer(self)  # type: ignore[arg-type]
        self.baseline_finish_timer.setTimerType(Qt.PreciseTimer)
        self.baseline_finish_timer.setSingleShot(True)
        self.baseline_finish_timer.timeout.connect(self._baseline_finish_test)
        self.baseline_countdown_timer = QTimer(self)  # type: ignore[arg-type]
        self.baseline_countdown_timer.setTimerType(Qt.PreciseTimer)
        self.baseline_countdown_timer.timeout.connect(self._baseline_update_countdown)

    def init_baseline_tab(self) -> None:
        if not PYQTGRAPH_AVAILABLE:
            layout = QVBoxLayout(self.baseline_tab)
            lbl = QLabel("基线稳定测试不可用，缺少 pyqtgraph 库。")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setFont(QFont("Microsoft YaHei", 14))
            layout.addWidget(lbl)
            return

        main_layout = QVBoxLayout(self.baseline_tab)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(12)

        top_layout = QHBoxLayout()
        top_layout.setSpacing(12)
        self._baseline_create_voltage_group(top_layout)
        self._baseline_create_live_status_group(top_layout)
        main_layout.addLayout(top_layout, stretch=3)

        bottom_layout = QHBoxLayout()
        bottom_layout.setSpacing(12)
        self._baseline_create_control_group(bottom_layout)
        self._baseline_create_results_group(bottom_layout)
        main_layout.addLayout(bottom_layout, stretch=2)

    def _baseline_create_voltage_group(self, parent_layout) -> None:
        group = QGroupBox("采样电压表")
        layout = QVBoxLayout(group)
        self.baseline_voltage_plot = pg.PlotWidget()
        self.baseline_voltage_plot.setBackground("w")
        self.baseline_voltage_plot.showGrid(x=True, y=True, alpha=0.3)
        self.baseline_voltage_plot.setLabel("left", "电压", units="V")
        self.baseline_voltage_plot.setLabel("bottom", "样本序号")
        self.baseline_voltage_curve = self.baseline_voltage_plot.plot(
            pen=pg.mkPen(color=COLOR_PRIMARY, width=2)
        )
        layout.addWidget(self.baseline_voltage_plot)
        parent_layout.addWidget(group, stretch=4)

    def _baseline_create_live_status_group(self, parent_layout) -> None:
        group = QGroupBox("实时状态")
        layout = QGridLayout(group)
        layout.setSpacing(10)
        self.baseline_current_voltage_value = self._baseline_create_live_value("0.0000 V", COLOR_PRIMARY)
        self.baseline_sample_count_value = self._baseline_create_live_value("0", COLOR_ACCENT_GREEN)
        self.baseline_remaining_value = self._baseline_create_live_value("00:00:00", COLOR_ACCENT_ORANGE)
        self.baseline_status_value = QLabel("未开始")
        self.baseline_status_value.setWordWrap(True)
        self.baseline_status_value.setStyleSheet(f"font-size: 14px; color: {COLOR_TEXT_SECONDARY};")

        layout.addWidget(QLabel("当前电压"), 0, 0)
        layout.addWidget(self.baseline_current_voltage_value, 0, 1)
        layout.addWidget(QLabel("有效样本"), 1, 0)
        layout.addWidget(self.baseline_sample_count_value, 1, 1)
        layout.addWidget(QLabel("剩余时间"), 2, 0)
        layout.addWidget(self.baseline_remaining_value, 2, 1)
        layout.addWidget(QLabel("状态"), 3, 0)
        layout.addWidget(self.baseline_status_value, 3, 1)
        self.baseline_progress_label = QLabel("进度")
        self.baseline_progress_value = QLabel("0%")
        self.baseline_progress_value.setStyleSheet(f"font-weight: 600; color: {COLOR_TEXT_PRIMARY};")
        layout.addWidget(self.baseline_progress_label, 4, 0)
        layout.addWidget(self.baseline_progress_value, 4, 1)
        parent_layout.addWidget(group, stretch=1)

    def _baseline_create_live_value(self, text: str, color: str) -> QLabel:
        label = QLabel(text)
        label.setStyleSheet(f"font-size: 24px; font-weight: 700; color: {color};")
        return label

    def _baseline_create_control_group(self, parent_layout) -> None:
        group = QGroupBox("操作区")
        layout = QVBoxLayout(group)
        layout.setSpacing(10)
        form = QFormLayout()
        form.setSpacing(8)

        min_duration, max_duration = SPECTRO_BASELINE_DURATION_RANGE
        self.baseline_duration_spin = QSpinBox()
        self.baseline_duration_spin.setRange(min_duration, max_duration)
        self.baseline_duration_spin.setValue(DEFAULT_BASELINE_DURATION_MIN)
        self.baseline_duration_spin.setSuffix(" min")

        min_warmup, max_warmup = SPECTRO_BASELINE_WARMUP_RANGE
        self.baseline_warmup_spin = QSpinBox()
        self.baseline_warmup_spin.setRange(min_warmup, max_warmup)
        self.baseline_warmup_spin.setValue(DEFAULT_BASELINE_WARMUP_S)
        self.baseline_warmup_spin.setSuffix(" s")

        min_points, max_points = SPECTRO_CHART_POINTS_RANGE
        self.baseline_chart_points_spin = QSpinBox()
        self.baseline_chart_points_spin.setRange(min_points, max_points)
        self.baseline_chart_points_spin.setSingleStep(SPECTRO_CHART_POINTS_STEP)
        self.baseline_chart_points_spin.setValue(self.baseline_max_data_points)
        self.baseline_chart_points_spin.setSuffix(" 点")
        self.baseline_chart_points_spin.valueChanged.connect(self._baseline_set_chart_points)

        # 兼容 settings_mixin 中既有的 spectro_baseline_* 设置键读取。
        self.spectro_baseline_duration_spin = self.baseline_duration_spin
        self.spectro_baseline_warmup_spin = self.baseline_warmup_spin

        form.addRow("测试时长:", self.baseline_duration_spin)
        form.addRow("预热丢弃:", self.baseline_warmup_spin)
        form.addRow("图表点数:", self.baseline_chart_points_spin)
        layout.addLayout(form)

        button_layout = QHBoxLayout()
        self.baseline_start_btn = QPushButton("开始测试")
        self.baseline_start_btn.setStyleSheet(BUTTON_SUCCESS)
        self.baseline_start_btn.clicked.connect(self._baseline_toggle_test)
        self.baseline_export_btn = QPushButton("导出 XLSX")
        self.baseline_export_btn.setStyleSheet(BUTTON_SECONDARY)
        self.baseline_export_btn.setEnabled(False)
        self.baseline_export_btn.clicked.connect(self._baseline_export_xlsx)
        button_layout.addWidget(self.baseline_start_btn)
        button_layout.addWidget(self.baseline_export_btn)
        layout.addLayout(button_layout)

        self.baseline_hint_label = QLabel("测试将复用分光 ADS 数据流；未采集时会自动启动 ADS。")
        self.baseline_hint_label.setWordWrap(True)
        self.baseline_hint_label.setStyleSheet(f"color: {COLOR_TEXT_MUTED};")
        layout.addWidget(self.baseline_hint_label)
        layout.addStretch(1)
        parent_layout.addWidget(group, stretch=1)

    def _baseline_create_results_group(self, parent_layout) -> None:
        group = QGroupBox("测试结果")
        layout = QGridLayout(group)
        layout.setSpacing(10)
        self.baseline_result_cards: dict[str, QLabel] = {}
        cards = [
            ("drift_v", "漂移", "V"),
            ("drift_percent", "漂移比例", "%"),
            ("peak_to_peak_v", "峰峰值", "V"),
            ("std_dev_v", "标准差", "V"),
            ("slope_v_per_min", "斜率", "V/min"),
            ("detrended_rms_v", "去趋势RMS", "V"),
            ("start_end_voltage", "起止均值", "V"),
            ("duration_s", "有效时长", "min"),
        ]
        for index, (key, title, unit) in enumerate(cards):
            self._baseline_add_metric_card(layout, index // 4, index % 4, key, title, unit)
        parent_layout.addWidget(group, stretch=2)

    def _baseline_add_metric_card(
        self,
        layout: QGridLayout,
        row: int,
        column: int,
        key: str,
        title: str,
        unit: str,
    ) -> None:
        card = QFrame()
        card.setStyleSheet(
            f"QFrame {{ background-color: {COLOR_BG_SUBTLE}; border: 1px solid {COLOR_BORDER_LIGHT}; "
            "border-radius: 8px; padding: 8px; }"
        )
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(4)
        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: 13px; color: {COLOR_TEXT_SECONDARY};")
        value_label = QLabel("--")
        value_label.setStyleSheet(f"font-size: 19px; font-weight: 700; color: {COLOR_TEXT_PRIMARY};")
        unit_label = QLabel(unit)
        unit_label.setStyleSheet(f"font-size: 12px; color: {COLOR_TEXT_MUTED};")
        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)
        card_layout.addWidget(unit_label)
        layout.addWidget(card, row, column)
        self.baseline_result_cards[key] = value_label

    def _baseline_set_chart_points(self, value: int, sync_spectro: bool = True) -> None:
        self.baseline_max_data_points = value
        if sync_spectro:
            self.spectro_max_data_points = value
            if hasattr(self, "spectro_chart_points_spin") and self.spectro_chart_points_spin.value() != value:
                self.spectro_chart_points_spin.blockSignals(True)
                self.spectro_chart_points_spin.setValue(value)
                self.spectro_chart_points_spin.blockSignals(False)
            if hasattr(self, "_spectro_trim_chart_data"):
                self._spectro_trim_chart_data()
                self._spectro_update_charts()
        self._baseline_trim_chart_data()
        self._baseline_update_chart()

    def _baseline_trim_chart_data(self) -> None:
        if len(self.baseline_voltage_data) > self.baseline_max_data_points:
            del self.baseline_voltage_data[:-self.baseline_max_data_points]

    def _baseline_toggle_test(self) -> None:
        if self.baseline_is_running:
            self._baseline_finish_test(manual_stop=True)
        else:
            self._baseline_start_test()

    def _baseline_start_test(self) -> None:
        owned_measurement = not self.spectro_is_measuring
        if owned_measurement and not self._spectro_start_measurement():
            return

        self.baseline_records.clear()
        self.baseline_voltage_data.clear()
        self.baseline_metrics = None
        self.baseline_is_running = True
        self.baseline_owned_measurement = owned_measurement
        self.baseline_start_time = time.time()
        duration_s = self.baseline_duration_spin.value() * 60
        self.baseline_end_time = self.baseline_start_time + duration_s
        self.baseline_started_at = datetime.now()
        self.baseline_finished_at = None
        self._baseline_set_controls_enabled(False)
        self._baseline_reset_result_cards()
        self.baseline_export_btn.setEnabled(False)
        self.baseline_start_btn.setText("停止测试")
        self.baseline_start_btn.setStyleSheet(BUTTON_DANGER)
        self.baseline_status_value.setText("测试中...")
        self.baseline_sample_count_value.setText("0")
        self.baseline_current_voltage_value.setText("0.0000 V")
        self.baseline_finish_timer.start(duration_s * 1000)
        self.baseline_countdown_timer.start(1000)
        self._baseline_update_countdown()
        self._baseline_update_chart()
        self.log(
            "基线稳定测试开始: "
            f"时长 {self.baseline_duration_spin.value()} min, "
            f"预热 {self.baseline_warmup_spin.value()} s"
        )

    def _baseline_finish_test(
        self,
        manual_stop: bool = False,
        stop_owned_measurement: bool = True,
    ) -> None:
        if not self.baseline_is_running:
            return

        if self.baseline_finish_timer.isActive():
            self.baseline_finish_timer.stop()
        if self.baseline_countdown_timer.isActive():
            self.baseline_countdown_timer.stop()

        owned_measurement = self.baseline_owned_measurement
        self.baseline_is_running = False
        self.baseline_owned_measurement = False
        self.baseline_finished_at = datetime.now()
        self._baseline_set_controls_enabled(True)
        self.baseline_start_btn.setText("开始测试")
        self.baseline_start_btn.setStyleSheet(BUTTON_SUCCESS)
        self.baseline_remaining_value.setText("00:00:00")
        self.baseline_progress_value.setText("100%" if not manual_stop else "已停止")

        valid_samples = [
            BaselineSample(timestamp_s=record.elapsed_s, voltage=record.voltage_v)
            for record in self.baseline_records
            if record.valid
        ]
        try:
            self.baseline_metrics = analyze_baseline(
                valid_samples,
                warmup_s=float(self.baseline_warmup_spin.value()),
            )
        except ValueError as exc:
            self.baseline_metrics = None
            self.baseline_status_value.setText("样本不足")
            self._baseline_reset_result_cards()
            self.log(f"基线稳定测试无法生成结果: {exc}")
        else:
            self._baseline_update_result_cards(self.baseline_metrics)
            self.baseline_status_value.setText("已手动停止" if manual_stop else "已完成")
            self.log(
                "基线稳定测试完成: "
                f"样本 {self.baseline_metrics.sample_count} 点, "
                f"漂移 {self.baseline_metrics.drift_v:+.6f} V, "
                f"斜率 {self.baseline_metrics.slope_v_per_min:+.6f} V/min"
            )

        self.baseline_export_btn.setEnabled(bool(self.baseline_records))
        if stop_owned_measurement and owned_measurement:
            self._spectro_stop_measurement()

    def _baseline_set_controls_enabled(self, enabled: bool) -> None:
        self.baseline_duration_spin.setEnabled(enabled)
        self.baseline_warmup_spin.setEnabled(enabled)

    def handle_baseline_packet(self, packet: dict, voltage: float, status: int) -> None:
        if not self.baseline_is_running:
            return

        elapsed_s = time.time() - self.baseline_start_time
        valid = self._baseline_is_valid_packet(status)
        record = BaselineExportRecord(
            elapsed_s=elapsed_s,
            voltage_v=voltage,
            raw_code=int(packet.get("raw_code", 0)),
            status=status,
            tca_channel=int(packet.get("tca_channel", 0)),
            valid=valid,
        )
        self.baseline_records.append(record)
        self.baseline_current_voltage_value.setText(f"{voltage:.4f} V")

        valid_count = sum(1 for item in self.baseline_records if item.valid)
        self.baseline_sample_count_value.setText(str(valid_count))
        if valid:
            self.baseline_voltage_data.append(voltage)
            self._baseline_trim_chart_data()
            self._baseline_update_chart()
            self.baseline_status_value.setText(f"测试中... 有效样本 {valid_count} 点")
        else:
            self.baseline_status_value.setText("测试中... 收到无效/异常分光包")

    def _baseline_update_chart(self) -> None:
        if hasattr(self, "baseline_voltage_curve"):
            self.baseline_voltage_curve.setData(self.baseline_voltage_data)

    def _baseline_update_countdown(self) -> None:
        if not self.baseline_is_running:
            return
        remaining_s = max(0, int(self.baseline_end_time - time.time()))
        self.baseline_remaining_value.setText(self._baseline_format_duration(remaining_s))
        total_s = max(1, self.baseline_duration_spin.value() * 60)
        progress = min(100, max(0, int((total_s - remaining_s) / total_s * 100)))
        self.baseline_progress_value.setText(f"{progress}%")

    def _baseline_format_duration(self, seconds: int) -> str:
        hours, rem = divmod(max(0, seconds), 3600)
        minutes, secs = divmod(rem, 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

    def _baseline_is_valid_packet(self, status: int) -> bool:
        return bool(status & 0x01) and not bool(status & 0x02) and not bool(status & 0x08)

    def _baseline_reset_result_cards(self) -> None:
        for label in self.baseline_result_cards.values():
            label.setText("--")

    def _baseline_update_result_cards(self, metrics: BaselineMetrics) -> None:
        values = {
            "drift_v": f"{metrics.drift_v:+.6f}",
            "drift_percent": f"{metrics.drift_percent:+.3f}",
            "peak_to_peak_v": f"{metrics.peak_to_peak_v:.6f}",
            "std_dev_v": f"{metrics.std_dev_v:.6f}",
            "slope_v_per_min": f"{metrics.slope_v_per_min:+.6f}",
            "detrended_rms_v": f"{metrics.detrended_rms_v:.6f}",
            "start_end_voltage": f"{metrics.start_voltage_v:.6f} -> {metrics.end_voltage_v:.6f}",
            "duration_s": f"{metrics.duration_s / 60.0:.2f}",
        }
        for key, text in values.items():
            self.baseline_result_cards[key].setText(text)

    def _baseline_export_xlsx(self) -> None:
        if not self.baseline_records:
            QMessageBox.warning(self, "导出错误", "没有基线测试数据可以导出")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            QMessageBox.critical(self, "导出错误", f"缺少 openpyxl 依赖: {exc}")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"baseline_stability_{ts}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "导出基线测试数据", filename, "Excel Files (*.xlsx)")
        if not path:
            return

        tables = build_baseline_export_tables(
            records=self.baseline_records,
            metrics=self.baseline_metrics,
            duration_min=self.baseline_duration_spin.value(),
            warmup_s=self.baseline_warmup_spin.value(),
            started_at=self.baseline_started_at.strftime("%Y-%m-%d %H:%M:%S") if self.baseline_started_at else "",
            finished_at=self.baseline_finished_at.strftime("%Y-%m-%d %H:%M:%S") if self.baseline_finished_at else "",
        )

        try:
            wb = Workbook()
            summary_ws = wb.active
            summary_ws.title = "summary"
            samples_ws = wb.create_sheet("samples")
            for row in tables.summary:
                summary_ws.append(row)
            for row in tables.samples:
                samples_ws.append(row)
            self._baseline_format_worksheet(summary_ws, get_column_letter, Font, PatternFill, Alignment)
            self._baseline_format_worksheet(samples_ws, get_column_letter, Font, PatternFill, Alignment)
            wb.save(path)
            self.log(f"基线测试数据已导出至 {os.path.basename(path)}")
        except Exception as exc:
            self.log(f"导出基线测试数据失败: {exc}")
            QMessageBox.critical(self, "导出错误", f"文件导出失败: {exc}")

    def _baseline_format_worksheet(self, ws, get_column_letter, Font, PatternFill, Alignment) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="DDEBFF")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 32)
