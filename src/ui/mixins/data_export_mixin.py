"""数据导出 Mixin 模块。

该模块提供数据导出和导入相关功能，包括：
- PID优化数据导出
- PID分析报告导出
- PID控制数据导出(Excel)
- 图表数据导出/导入

Note:
    此模块设计为 Mixin 类，需要与 QMainWindow 子类一起使用。
"""

import csv
from datetime import datetime
from typing import Optional

from PySide6.QtCore import QPointF
from PySide6.QtWidgets import (
    QButtonGroup,
    QDialog,
    QFileDialog,
    QLabel,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QVBoxLayout,
)


class DataExportMixin:
    """数据导出功能 Mixin。

    提供PID数据导出、报告生成和图表数据导入导出功能。

    Attributes:
        pid_optimizer: PID优化器实例
        pid_analyzer: PID分析器实例
        chart_view: 图表视图
    """

    def _export_pid_optimization_data(self) -> None:
        """导出PID优化数据到CSV文件。"""
        history = self.pid_optimizer.get_history_summary()
        if not history:
            QMessageBox.information(self, "提示", "没有可导出的优化数据")
            return

        default_name = f"pid_optimization_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出PID优化数据", default_name, "CSV文件 (*.csv)"
        )

        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "迭代", "Kp", "Ki", "Kd", "原始得分", "调整得分",
                    "最大过冲°", "收敛时间ms", "RSD%"
                ])
                for record in history:
                    writer.writerow([
                        record.get("index", 0),
                        f"{record.get('Kp', 0):.4f}",
                        f"{record.get('Ki', 0):.5f}",
                        f"{record.get('Kd', 0):.4f}",
                        f"{record.get('avg_score', 0):.1f}",
                        f"{record.get('adjusted_score', record.get('avg_score', 0)):.1f}",
                        f"{record.get('max_overshoot', 0):.2f}",
                        f"{record.get('avg_conv_time', 0):.0f}",
                        f"{record.get('convergence_rsd', 0):.1f}",
                    ])

                writer.writerow([])
                writer.writerow(["最优参数（贝叶斯优化 + 非线性惩罚）"])
                if self.pid_optimizer.best_params:
                    writer.writerow(["Kp", f"{self.pid_optimizer.best_params.Kp:.4f}"])
                    writer.writerow(["Ki", f"{self.pid_optimizer.best_params.Ki:.5f}"])
                    writer.writerow(["Kd", f"{self.pid_optimizer.best_params.Kd:.4f}"])
                    writer.writerow(["最优得分（惩罚后）", f"{self.pid_optimizer.best_score:.1f}"])

            self.log(f"[导出] PID优化数据已导出到: {file_path}")
            QMessageBox.information(self, "成功", f"数据已导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {e}")

    def export_pid_report(self) -> None:
        """导出 PID 分析报告。"""
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装 pandas 库：pip install pandas")
            return

        report_data = []
        for motor in ["X", "Y", "Z", "A"]:
            stats = self.pid_analyzer.get_stats_summary(motor)
            stats["motor"] = motor
            report_data.append(stats)

        if not any(d["total_runs"] > 0 for d in report_data):
            QMessageBox.warning(self, "警告", "没有可导出的 PID 运行数据")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pid_report_{timestamp}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "导出 PID 报告", filename, "Excel文件 (*.xlsx)")

        if path:
            try:
                df = pd.DataFrame(report_data)
                df = df[[
                    "motor", "total_runs", "success_rate", "avg_convergence_time",
                    "min_convergence_time", "max_convergence_time", "avg_final_error",
                    "max_final_error", "timeout_count", "fail_count"
                ]]
                df.columns = [
                    "电机", "运行次数", "成功率", "平均收敛时间", "最短收敛时间",
                    "最长收敛时间", "平均最终误差", "最大最终误差", "超时次数", "失败次数"
                ]
                df.to_excel(path, index=False)
                self.log(f"PID 报告已导出至 {path}")
            except Exception as e:
                QMessageBox.critical(self, "导出错误", f"文件写入失败: {str(e)}")

    def export_pid_data(self) -> None:
        """导出 PID 控制数据到 Excel 文件。"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装 openpyxl 库：pip install openpyxl")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pid_data_{timestamp}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "导出 PID 数据", filename, "Excel 文件 (*.xlsx)")

        if not path:
            return

        try:
            wb = Workbook()

            # 样式定义
            header_font = Font(name="Arial", size=11, bold=True)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Arial", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            def apply_header_style(ws, row=1):
                for cell in ws[row]:
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

            def apply_data_style(ws, start_row=2):
                for row in ws.iter_rows(min_row=start_row):
                    for cell in row:
                        cell.font = data_font
                        cell.alignment = center_align
                        cell.border = thin_border

            def auto_column_width(ws, min_width=10, max_width=25):
                for column_cells in ws.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max(max_length + 2, min_width), max_width)
                    ws.column_dimensions[column].width = adjusted_width

            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"
            headers = ["Motor", "Total Runs", "Success Rate", "Avg Conv Time (s)",
                      "Min Conv Time (s)", "Max Conv Time (s)", "Avg Error (°)",
                      "Max Error (°)", "Timeout", "Failed"]
            ws_summary.append(headers)

            for motor in ["X", "Y", "Z", "A"]:
                stats = self.pid_analyzer.stats[motor]
                row = [
                    motor, stats.total_runs, f"{stats.success_rate:.1f}%",
                    f"{stats.avg_convergence_time:.3f}" if stats.successful_runs > 0 else "-",
                    f"{stats.min_convergence_time:.3f}" if stats.min_convergence_time else "-",
                    f"{stats.max_convergence_time:.3f}" if stats.max_convergence_time else "-",
                    f"{stats.avg_final_error:.3f}" if stats.successful_runs > 0 else "-",
                    f"{stats.max_final_error:.3f}" if stats.max_final_error > 0 else "-",
                    stats.timeout_runs, stats.failed_runs
                ]
                ws_summary.append(row)

            apply_header_style(ws_summary)
            apply_data_style(ws_summary)
            auto_column_width(ws_summary)
            ws_summary.freeze_panes = "A2"

            # Sheet 2: Error Distribution
            ws_error_dist = wb.create_sheet("Error Distribution")
            ws_error_dist.append(["Index", "Motor", "Final Error (°)"])
            idx = 1
            for motor in ["X", "Y", "Z", "A"]:
                for error in self.pid_analyzer.stats[motor].error_distribution:
                    ws_error_dist.append([idx, motor, f"{error:.4f}"])
                    idx += 1
            apply_header_style(ws_error_dist)
            apply_data_style(ws_error_dist)
            auto_column_width(ws_error_dist)
            ws_error_dist.freeze_panes = "A2"

            # Sheet 3: Run History
            ws_history = wb.create_sheet("Run History")
            ws_history.append(["Run ID", "Motor", "Target (°)", "Precision (°)",
                              "Duration (s)", "Final Angle (°)", "Final Error (°)", "Status"])
            run_id = 1
            for motor in ["X", "Y", "Z", "A"]:
                for record in self.pid_analyzer.history[motor]:
                    row = [
                        run_id, motor, f"{record.target_angle:.2f}", f"{record.precision:.2f}",
                        f"{record.duration:.3f}" if record.duration else "-",
                        f"{record.final_angle:.3f}" if record.final_angle is not None else "-",
                        f"{record.final_error:.3f}" if record.final_error is not None else "-",
                        record.status.value
                    ]
                    ws_history.append(row)
                    run_id += 1
            apply_header_style(ws_history)
            apply_data_style(ws_history)
            auto_column_width(ws_history)
            ws_history.freeze_panes = "A2"

            # Sheet 4-7: Realtime data sheets
            self._export_realtime_sheets(wb, apply_header_style, apply_data_style, auto_column_width)

            wb.save(path)
            self.log(f"PID 数据已导出至 {path}")

        except Exception as e:
            QMessageBox.critical(self, "导出错误", f"数据导出失败: {str(e)}")

    def _export_realtime_sheets(self, wb, apply_header_style, apply_data_style, auto_column_width):
        """导出实时数据表格（内部辅助方法）。"""
        # Sheet 4: Realtime Position
        ws_position = wb.create_sheet("Realtime Position")
        pos_headers = ["Time (s)"]
        for motor in ["X", "Y", "Z", "A"]:
            pos_headers.extend([f"{motor}_Target", f"{motor}_Actual", f"{motor}_Theo"])
        ws_position.append(pos_headers)

        all_times = set()
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_position_data(motor):
                all_times.add(round(data[0], 3))

        pos_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_position_data(motor):
                t = round(data[0], 3)
                pos_dict[motor][t] = (data[1], data[2], data[3])

        for t in sorted(all_times):
            row = [f"{t:.3f}"]
            for motor in ["X", "Y", "Z", "A"]:
                if t in pos_dict[motor]:
                    target, actual, theo = pos_dict[motor][t]
                    row.extend([f"{target:.3f}", f"{actual:.3f}", f"{theo:.3f}"])
                else:
                    row.extend(["-", "-", "-"])
            ws_position.append(row)

        apply_header_style(ws_position)
        apply_data_style(ws_position)
        auto_column_width(ws_position)
        ws_position.freeze_panes = "A2"

        # Sheet 5: Realtime Output
        ws_output = wb.create_sheet("Realtime Output")
        ws_output.append(["Time (s)", "X_Output (RPM)", "Y_Output (RPM)", "Z_Output (RPM)", "A_Output (RPM)"])

        all_times = set()
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_output_data(motor):
                all_times.add(round(data[0], 3))

        out_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_output_data(motor):
                out_dict[motor][round(data[0], 3)] = data[1]

        for t in sorted(all_times):
            row = [f"{t:.3f}"]
            for motor in ["X", "Y", "Z", "A"]:
                row.append(f"{out_dict[motor][t]:.3f}" if t in out_dict[motor] else "-")
            ws_output.append(row)

        apply_header_style(ws_output)
        apply_data_style(ws_output)
        auto_column_width(ws_output)
        ws_output.freeze_panes = "A2"

        # Sheet 6: Realtime Error
        ws_error = wb.create_sheet("Realtime Error")
        ws_error.append(["Time (s)", "X_Error (°)", "Y_Error (°)", "Z_Error (°)", "A_Error (°)"])

        all_times = set()
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_error_data(motor):
                all_times.add(round(data[0], 3))

        err_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_error_data(motor):
                err_dict[motor][round(data[0], 3)] = data[1]

        for t in sorted(all_times):
            row = [f"{t:.3f}"]
            for motor in ["X", "Y", "Z", "A"]:
                row.append(f"{err_dict[motor][t]:.3f}" if t in err_dict[motor] else "-")
            ws_error.append(row)

        apply_header_style(ws_error)
        apply_data_style(ws_error)
        auto_column_width(ws_error)
        ws_error.freeze_panes = "A2"

        # Sheet 7: Realtime Load
        ws_load = wb.create_sheet("Realtime Load")
        ws_load.append(["Time (s)", "X_Load (°)", "Y_Load (°)", "Z_Load (°)", "A_Load (°)"])

        all_times = set()
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_load_data(motor):
                all_times.add(round(data[0], 3))

        load_dict = {motor: {} for motor in ["X", "Y", "Z", "A"]}
        for motor in ["X", "Y", "Z", "A"]:
            for data in self.pid_analyzer.get_export_load_data(motor):
                load_dict[motor][round(data[0], 3)] = data[1]

        for t in sorted(all_times):
            row = [f"{t:.3f}"]
            for motor in ["X", "Y", "Z", "A"]:
                row.append(f"{load_dict[motor][t]:.3f}" if t in load_dict[motor] else "-")
            ws_load.append(row)

        apply_header_style(ws_load)
        apply_data_style(ws_load)
        auto_column_width(ws_load)
        ws_load.freeze_panes = "A2"

    def save_chart_image(self) -> None:
        """保存图表为图片文件。"""
        options = QFileDialog.Options()
        path, _ = QFileDialog.getSaveFileName(
            self, "保存图表", "", "PNG图像 (*.png);;JPEG图像 (*.jpg)", options=options
        )

        if path:
            pixmap = self.chart_view.grab()
            if pixmap.save(path):
                self.log(f"图表已保存至 {path}")
            else:
                QMessageBox.warning(self, "错误", "图片保存失败")

    def export_chart_data(self) -> None:
        """导出图表数据到Excel。"""
        QMessageBox.information(self, "提示", "图表导出功能已弃用")
        self.log("图表导出功能已弃用")
        return
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装pandas库：pip install pandas")
            return

        selected = self.export_motor_combo.currentText()
        motor_map = {"全部": ["X", "Y", "Z", "A"], "X轴": ["X"], "Y轴": ["Y"], "Z轴": ["Z"], "A轴": ["A"]}
        selected_motors = motor_map[selected]

        chart_data = self.chart_view.chart().get_chart_data()
        max_length = max(len(chart_data[m]) for m in selected_motors if chart_data[m]) if any(chart_data.values()) else 0
        if max_length == 0:
            QMessageBox.warning(self, "警告", "当前没有可导出的数据")
            return

        data_dict = {}
        for motor in selected_motors:
            values = chart_data[motor]
            if len(values) < max_length:
                values += [float("nan")] * (max_length - len(values))
            data_dict[f"{motor}轴偏差"] = values

        options = QFileDialog.Options()
        path, _ = QFileDialog.getSaveFileName(self, "导出数据", "", "Excel文件 (*.xlsx)", options=options)

        if path:
            try:
                df = pd.DataFrame(data_dict)
                df.index.name = "采样序列"
                with pd.ExcelWriter(path) as writer:
                    df.to_excel(writer)
                self.log(f"数据已导出至 {path}")
            except Exception as e:
                QMessageBox.critical(self, "导出错误", f"文件写入失败: {str(e)}")

    def import_chart_data(self) -> None:
        """导入图表数据。"""
        QMessageBox.information(self, "提示", "图表导入功能已弃用")
        self.log("图表导入功能已弃用")
        return
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装pandas库：pip install pandas")
            return

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", "", "Excel文件 (*.xlsx);;CSV文件 (*.csv)", options=options
        )
        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx"):
                df = pd.read_excel(file_path, index_col=0)
            else:
                df = pd.read_csv(file_path, index_col=0)

            valid_columns = []
            motor_mapping = {"X轴偏差": "X", "Y轴偏差": "Y", "Z轴偏差": "Z", "A轴偏差": "A"}

            for col in df.columns:
                if col in motor_mapping:
                    valid_columns.append(col)

            if not valid_columns:
                raise ValueError("未检测到有效偏差数据列（列名示例：X轴偏差）")

            mode = self._show_import_dialog()
            if not mode:
                return

            for col in valid_columns:
                motor = motor_mapping[col]
                data = df[col].dropna().tolist()

                if mode == "replace":
                    self.chart_view.chart().data[motor].clear()
                    self.chart_view.chart().data[motor].extend(data)
                else:
                    self.chart_view.chart().data[motor].extend(data)

                points = [QPointF(x, y) for x, y in enumerate(self.chart_view.chart().data[motor])]
                self.chart_view.chart().series[motor].replace(points[-200:])

            self.chart_view.chart().auto_scale_axes()
            self._update_stats_after_import(df, valid_columns)
            self.log(f"成功导入{len(valid_columns)}轴数据")
        except Exception as e:
            QMessageBox.critical(self, "导入错误", f"数据导入失败: {str(e)}")

    def _show_import_dialog(self) -> Optional[str]:
        """显示导入选项对话框。"""
        dialog = QDialog(self)
        dialog.setWindowTitle("导入选项")
        layout = QVBoxLayout(dialog)

        mode_group = QButtonGroup(dialog)
        rb_replace = QRadioButton("替换现有数据", dialog)
        rb_append = QRadioButton("追加到现有数据", dialog)
        rb_replace.setChecked(True)

        mode_group.addButton(rb_replace)
        mode_group.addButton(rb_append)

        btn_confirm = QPushButton("确认导入", dialog)
        btn_confirm.clicked.connect(dialog.accept)

        layout.addWidget(QLabel("请选择数据导入模式:"))
        layout.addWidget(rb_replace)
        layout.addWidget(rb_append)
        layout.addWidget(btn_confirm)

        if dialog.exec() == QDialog.Accepted:
            return "replace" if rb_replace.isChecked() else "append"
        return None

    def _update_stats_after_import(self, df, valid_columns) -> None:
        """更新统计信息（仅更新导入的轴）。"""
        for col in valid_columns:
            motor = col[0]
            data = df[col].dropna()

            if not data.empty:
                labels = self.stats_widgets[motor]
                labels["current"].setText(f"{data.iloc[-1]:.2f}°")
                labels["average"].setText(f"{data.mean():.2f}°")
                labels["dev_rate"].setText(
                    f"{(data.iloc[-1] / data.mean() * 100 if data.mean() != 0 else 0):.1f}%"
                )

    def _update_all_stats(self, df) -> None:
        """更新所有统计信息。"""
        self.cumulative_deviations = {m: 0.0 for m in ["X", "Y", "Z", "A"]}

        for motor in ["X", "Y", "Z", "A"]:
            col = f"{motor}轴偏差"
            if col not in df.columns:
                continue
            data = df[col].dropna()

            if not data.empty:
                current = data.iloc[-1]
                avg = data.mean()

                labels = self.stats_widgets[motor]
                labels["current"].setText(f"{current:.2f}°")
                labels["average"].setText(f"{avg:.2f}°")
                labels["dev_rate"].setText(f"{(current / avg * 100 if avg != 0 else 0):.1f}%")
